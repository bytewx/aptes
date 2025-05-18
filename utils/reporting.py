#!/usr/bin/env python3
"""
Reporting utilities for APTES
"""

import json
import logging
import os
import threading
import time
from datetime import datetime, date

logger = logging.getLogger('aptes.reporting')

# Check for optional dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel reporting disabled")

class ReportGenerator:
    """Report generation utilities for APTES"""
    
    def __init__(self, results, target, output_dir="reports"):
        """
        Initialize the report generator
        
        Args:
            results (dict): Results data to be reported
            target (str): Target of the assessment
            output_dir (str): Directory to save reports
        """
        self.results = results
        self.target = target
        self.output_dir = output_dir
        
        # Risk colors for Excel reports
        if EXCEL_AVAILABLE:
            self.risk_colors = {
                "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                "high": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
                "medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                "low": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
                "info": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            }
    
    def generate_report(self, format="all", phase="preexploit"):
        """
        Generate report in specified format
        
        Args:
            format (str): Report format (json, excel, md, all)
            phase (str): Assessment phase to report on
        
        Returns:
            dict: Dictionary of generated report filenames
        """
        today = date.today().strftime("%Y%m%d")
        target_safe = self.target.replace(".", "_").replace(":", "_").replace("/", "_")
        base_filename = f"{self.output_dir}/{target_safe}_{phase}_{today}"
        
        report_files = {}
        
        if format in ["json", "all"]:
            json_file = f"{base_filename}.json"
            self.generate_json_report(json_file, phase)
            report_files["json"] = json_file
            
        if format in ["excel", "all"]:
            if EXCEL_AVAILABLE:
                excel_file = f"{base_filename}.xlsx"
                self.generate_excel_report(excel_file, phase)
                report_files["excel"] = excel_file
            else:
                logger.warning("Excel report skipped - openpyxl library not available")
                report_files["excel"] = None
            
        if format in ["md", "all"]:
            md_file = f"{base_filename}.md"
            self.generate_markdown_report(md_file, phase)
            report_files["markdown"] = md_file
            
        logger.info(f"Reports generated in {self.output_dir} directory")
        return report_files
    
    def generate_json_report(self, filename, phase="preexploit"):
        """
        Generate JSON report
        
        Args:
            filename (str): Output filename
            phase (str): Assessment phase to report on
        """
        data = self.results.get(phase, {})
        with open(filename, "w") as f:
            json.dump(data, f, indent=4)
    
    def generate_excel_report(self, filename, phase="preexploit"):
        """
        Generate Excel report
        
        Args:
            filename (str): Output filename
            phase (str): Assessment phase to report on
        """
        if not EXCEL_AVAILABLE:
            logger.error("Cannot generate Excel report - openpyxl library not available")
            return
        
        # Get phase data
        data = self.results.get(phase, {})
        
        wb = Workbook()
        
        # Summary sheet
        summary = wb.active
        summary.title = "Summary"
        summary['A1'] = f"{phase.capitalize()} Report"
        summary['A1'].font = Font(bold=True, size=14)
        summary['A3'] = "Target:"
        summary['B3'] = self.target
        summary['A4'] = "Timestamp:"
        summary['B4'] = data.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Phase-specific sheets
        if phase == "preexploit":
            self._add_preexploit_sheets(wb, data)
        elif phase == "recon":
            self._add_recon_sheets(wb, data)
        elif phase == "exploit":
            self._add_exploit_sheets(wb, data)
        elif phase == "postexploit":
            self._add_postexploit_sheets(wb, data)
        
        # Adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(filename)
    
    def _add_preexploit_sheets(self, workbook, data):
        """Add pre-exploitation sheets to Excel workbook"""
        # Vulnerabilities sheet
        if "vulnerability_validation" in data and "vulnerabilities" in data["vulnerability_validation"]:
            vulns_sheet = workbook.create_sheet("Vulnerabilities")
            vulns_sheet['A1'] = "Validated Vulnerabilities"
            vulns_sheet['A1'].font = Font(bold=True, size=14)
            
            vulns_sheet['A3'] = "Name"
            vulns_sheet['B3'] = "Host"
            vulns_sheet['C3'] = "Port"
            vulns_sheet['D3'] = "Service"
            vulns_sheet['E3'] = "Risk Level"
            vulns_sheet['F3'] = "CVE"
            vulns_sheet['G3'] = "Details"
            
            row = 4
            for vuln in data["vulnerability_validation"]["vulnerabilities"]:
                vulns_sheet[f'A{row}'] = vuln.get("vulnerability", "Unknown")
                vulns_sheet[f'B{row}'] = vuln.get("host", "Unknown")
                vulns_sheet[f'C{row}'] = vuln.get("port", "Unknown")
                vulns_sheet[f'D{row}'] = vuln.get("service", "Unknown")
                vulns_sheet[f'E{row}'] = vuln.get("risk_level", "Unknown")
                vulns_sheet[f'F{row}'] = vuln.get("cve", "N/A")
                vulns_sheet[f'G{row}'] = vuln.get("details", "No details available")
                
                risk = vuln.get("risk_level", "").lower()
                if risk in self.risk_colors:
                    vulns_sheet[f'E{row}'].fill = self.risk_colors[risk]
                
                row += 1
        
        # Web Findings sheet
        if "webapp_analysis" in data and "findings" in data["webapp_analysis"]:
            web_sheet = workbook.create_sheet("Web Findings")
            web_sheet['A1'] = "Web Application Findings"
            web_sheet['A1'].font = Font(bold=True, size=14)
            
            web_sheet['A3'] = "Finding"
            web_sheet['B3'] = "URL"
            web_sheet['C3'] = "Category"
            web_sheet['D3'] = "Risk Level"
            web_sheet['E3'] = "Description"
            web_sheet['F3'] = "Recommendation"
            
            row = 4
            for finding in data["webapp_analysis"]["findings"]:
                web_sheet[f'A{row}'] = finding.get("finding", "Unknown")
                host_port = f"{finding.get('host', '')}:{finding.get('port', '')}"
                web_sheet[f'B{row}'] = finding.get("url", host_port)
                web_sheet[f'C{row}'] = finding.get("category", "Unknown")
                web_sheet[f'D{row}'] = finding.get("risk_level", "Unknown")
                web_sheet[f'E{row}'] = finding.get("description", "No description available")
                web_sheet[f'F{row}'] = finding.get("recommendation", "No recommendation available")
                
                risk = finding.get("risk_level", "").lower()
                if risk in self.risk_colors:
                    web_sheet[f'D{row}'].fill = self.risk_colors[risk]
                
                row += 1
        
        # Attack Vectors sheet
        if "attack_vectors" in data:
            attack_sheet = workbook.create_sheet("Attack Vectors")
            attack_sheet['A1'] = "Attack Vectors"
            attack_sheet['A1'].font = Font(bold=True, size=14)
            
            attack_sheet['A3'] = "Vector"
            attack_sheet['B3'] = "Target"
            attack_sheet['C3'] = "Type"
            attack_sheet['D3'] = "Risk Level"
            attack_sheet['E3'] = "Description"
            
            row = 4
            for vector in data["attack_vectors"]:
                attack_sheet[f'A{row}'] = vector.get("name", "Unknown")
                attack_sheet[f'B{row}'] = vector.get("target", "Unknown")
                attack_sheet[f'C{row}'] = vector.get("type", "Unknown")
                attack_sheet[f'D{row}'] = vector.get("risk_level", "Unknown")
                attack_sheet[f'E{row}'] = vector.get("description", "No description available")
                
                risk = vector.get("risk_level", "").lower()
                if risk in self.risk_colors:
                    attack_sheet[f'D{row}'].fill = self.risk_colors[risk]
                
                row += 1
    
    def _add_recon_sheets(self, workbook, data):
        """Add reconnaissance sheets to Excel workbook"""
        # Passive recon sheet
        if "passive" in data:
            passive_sheet = workbook.create_sheet("Passive Recon")
            passive_sheet['A1'] = "Passive Reconnaissance Results"
            passive_sheet['A1'].font = Font(bold=True, size=14)
            
            # DNS Records
            row = 3
            if "dns" in data["passive"]:
                passive_sheet[f'A{row}'] = "DNS Records"
                passive_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for record_type, records in data["passive"]["dns"].items():
                    passive_sheet[f'A{row}'] = f"{record_type} Records:"
                    col = 'B'
                    for record in records:
                        passive_sheet[f'{col}{row}'] = str(record)
                        col = chr(ord(col) + 1)
                    row += 1
                
                row += 1
            
            # Subdomains
            if "subdomains" in data["passive"]:
                passive_sheet[f'A{row}'] = "Subdomains"
                passive_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for i, subdomain in enumerate(data["passive"]["subdomains"]):
                    passive_sheet[f'A{row}'] = subdomain
                    row += 1
                
                row += 1
            
            # SSL Info
            if "ssl_info" in data["passive"]:
                passive_sheet[f'A{row}'] = "SSL Certificate Information"
                passive_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ssl_info = data["passive"]["ssl_info"]
                for key, value in ssl_info.items():
                    passive_sheet[f'A{row}'] = key.capitalize()
                    passive_sheet[f'B{row}'] = str(value)
                    row += 1
        
        # Active recon sheet
        if "active" in data:
            active_sheet = workbook.create_sheet("Active Recon")
            active_sheet['A1'] = "Active Reconnaissance Results"
            active_sheet['A1'].font = Font(bold=True, size=14)
            
            # Port scan results
            row = 3
            if "ports" in data["active"]:
                active_sheet[f'A{row}'] = "Port Scan Results"
                active_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                active_sheet[f'A{row}'] = "Host"
                active_sheet[f'B{row}'] = "Protocol"
                active_sheet[f'C{row}'] = "Port"
                active_sheet[f'D{row}'] = "State"
                active_sheet[f'E{row}'] = "Service"
                row += 1
                
                for host, host_data in data["active"]["ports"].items():
                    for proto, proto_data in host_data.items():
                        for port, port_data in proto_data.items():
                            active_sheet[f'A{row}'] = host
                            active_sheet[f'B{row}'] = proto
                            active_sheet[f'C{row}'] = port
                            active_sheet[f'D{row}'] = port_data.get("state", "unknown")
                            active_sheet[f'E{row}'] = port_data.get("service", "unknown")
                            row += 1
                
                row += 1
        
        # Web crawler results sheet
        if "webcrawler" in data:
            crawler_sheet = workbook.create_sheet("Web Crawler")
            crawler_sheet['A1'] = "Web Crawler Results"
            crawler_sheet['A1'].font = Font(bold=True, size=14)
            
            row = 3
            crawler_sheet[f'A{row}'] = "Summary"
            crawler_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            crawler_sheet[f'A{row}'] = "Crawled URLs:"
            crawler_sheet[f'B{row}'] = data["webcrawler"].get("crawled_urls", 0)
            row += 1
            
            crawler_sheet[f'A{row}'] = "Forms Found:"
            crawler_sheet[f'B{row}'] = data["webcrawler"].get("forms_found", 0)
            row += 1
            
            crawler_sheet[f'A{row}'] = "Potential Vulnerable URLs:"
            crawler_sheet[f'B{row}'] = len(data["webcrawler"].get("potential_vulnerable_urls", []))
            row += 2
            
            # Display forms
            if "sitemap" in data["webcrawler"]:
                crawler_sheet[f'A{row}'] = "Forms"
                crawler_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                crawler_sheet[f'A{row}'] = "URL"
                crawler_sheet[f'B{row}'] = "Form Action"
                crawler_sheet[f'C{row}'] = "Method"
                crawler_sheet[f'D{row}'] = "Input Fields"
                row += 1
                
                for url, page_data in data["webcrawler"]["sitemap"].items():
                    if "forms" in page_data and page_data["forms"]:
                        for form in page_data["forms"]:
                            crawler_sheet[f'A{row}'] = form.get("form_url", url)
                            crawler_sheet[f'B{row}'] = form.get("action", "")
                            crawler_sheet[f'C{row}'] = form.get("method", "GET")
                            crawler_sheet[f'D{row}'] = ", ".join([f"{input.get('name', '')} ({input.get('type', '')})" 
                                                              for input in form.get("inputs", [])])
                            row += 1
                
    def _add_exploit_sheets(self, workbook, data):
        """Add exploitation sheets to Excel workbook"""
        # Exploitation summary sheet
        summary_sheet = workbook.create_sheet("Exploitation Summary")
        summary_sheet['A1'] = "Exploitation Results"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        row = 3
        if "exploitation_summary" in data:
            summary = data["exploitation_summary"]
            summary_sheet[f'A{row}'] = "Attempts:"
            summary_sheet[f'B{row}'] = summary.get("attempts", 0)
            row += 1
            
            summary_sheet[f'A{row}'] = "Successful:"
            summary_sheet[f'B{row}'] = summary.get("successful", 0)
            row += 1
            
            summary_sheet[f'A{row}'] = "Failed:"
            summary_sheet[f'B{row}'] = summary.get("failed", 0)
            row += 2
        
        # Exploits sheet
        if "exploits" in data:
            summary_sheet[f'A{row}'] = "Exploits Used"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            summary_sheet[f'A{row}'] = "Name"
            summary_sheet[f'B{row}'] = "Target"
            summary_sheet[f'C{row}'] = "Success"
            summary_sheet[f'D{row}'] = "Details"
            row += 1
            
            for exploit in data["exploits"]:
                summary_sheet[f'A{row}'] = exploit.get("name", "Unknown")
                summary_sheet[f'B{row}'] = exploit.get("target", "Unknown")
                summary_sheet[f'C{row}'] = "Yes" if exploit.get("success", False) else "No"
                summary_sheet[f'D{row}'] = exploit.get("details", "")
                row += 1
        
        # Shells sheet
        if "shells" in data and data["shells"]:
            shells_sheet = workbook.create_sheet("Shells")
            shells_sheet['A1'] = "Obtained Shells"
            shells_sheet['A1'].font = Font(bold=True, size=14)
            
            shells_sheet['A3'] = "Type"
            shells_sheet['B3'] = "Target"
            shells_sheet['C3'] = "Port"
            shells_sheet['D3'] = "Privileges"
            shells_sheet['E3'] = "Notes"
            
            row = 4
            for shell in data["shells"]:
                shells_sheet[f'A{row}'] = shell.get("type", "Unknown")
                shells_sheet[f'B{row}'] = shell.get("target", "Unknown")
                shells_sheet[f'C{row}'] = shell.get("port", "Unknown")
                shells_sheet[f'D{row}'] = shell.get("privileges", "Unknown")
                shells_sheet[f'E{row}'] = shell.get("notes", "")
                row += 1
    
    def _add_postexploit_sheets(self, workbook, data):
        """Add post-exploitation sheets to Excel workbook"""
        # Post-exploitation summary sheet
        summary_sheet = workbook.create_sheet("Post-Exploitation")
        summary_sheet['A1'] = "Post-Exploitation Results"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        # Persistence mechanisms
        row = 3
        if "persistence" in data and data["persistence"]:
            summary_sheet[f'A{row}'] = "Persistence Mechanisms"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            summary_sheet[f'A{row}'] = "Host"
            summary_sheet[f'B{row}'] = "Technique"
            summary_sheet[f'C{row}'] = "Location"
            summary_sheet[f'D{row}'] = "Notes"
            row += 1
            
            for mechanism in data["persistence"]:
                summary_sheet[f'A{row}'] = mechanism.get("host", "Unknown")
                summary_sheet[f'B{row}'] = mechanism.get("technique", "Unknown")
                summary_sheet[f'C{row}'] = mechanism.get("location", "Unknown")
                summary_sheet[f'D{row}'] = mechanism.get("notes", "")
                row += 1
            
            row += 1
        
        # Data exfiltration
        if "data_exfiltration" in data and data["data_exfiltration"]:
            summary_sheet[f'A{row}'] = "Data Exfiltration"
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            summary_sheet[f'A{row}'] = "Host"
            summary_sheet[f'B{row}'] = "Data Type"
            summary_sheet[f'C{row}'] = "Size"
            summary_sheet[f'D{row}'] = "Location"
            row += 1
            
            for exfil in data["data_exfiltration"]:
                summary_sheet[f'A{row}'] = exfil.get("host", "Unknown")
                summary_sheet[f'B{row}'] = ", ".join(exfil.get("data_types", ["Unknown"]))
                summary_sheet[f'C{row}'] = f"{exfil.get('total_size', 0) / (1024*1024):.2f} MB"
                summary_sheet[f'D{row}'] = exfil.get("location", "Unknown")
                row += 1

    def _render_llm_suggestions(self, llm_suggestions):
        """Render LLM suggestions as HTML."""
        if not llm_suggestions:
            return "<i>No LLM exploitation suggestions available.</i>"
        if isinstance(llm_suggestions, dict):
            if "raw_response" in llm_suggestions:
                content = llm_suggestions["raw_response"]
            else:
                import json
                content = "<pre class='json-block'>" + json.dumps(llm_suggestions, indent=2, ensure_ascii=False) + "</pre>"
        else:
            content = str(llm_suggestions)
        # If markdown-like, convert simple newlines and bold
        content = content.replace("\n", "<br>")
        content = content.replace("**", "<b>").replace("__", "<u>")
        return f"<div class='section'><h2>🤖 LLM Exploitation Suggestions</h2><div class='json-block'>{content}</div></div>"

    def generate_combined_html_report(self, filename=None):
        """
        Generate a combined HTML report for all phases.
        Args:
            filename (str): Output HTML filename.
        """
        from datetime import datetime

        # Prepare output filename
        today = datetime.now().strftime("%Y%m%d")
        target_safe = self.target.replace(".", "_").replace(":", "_").replace("/", "_")
        if not filename:
            filename = f"{self.output_dir}/{target_safe}_full_{today}.html"
        os.makedirs(self.output_dir, exist_ok=True)

        # CSS and helpers
        css = """
        <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #f6f8fa; color: #222; margin: 0; }
        .container { max-width: 1200px; margin: 30px auto; background: #fff; border-radius: 12px; box-shadow: 0 4px 24px #0001; padding: 32px; }
        h1, h2, h3 { color: #2d6cdf; margin-top: 1.5em; }
        h1 { border-bottom: 2px solid #eaecef; padding-bottom: 0.3em; }
        .section { margin-bottom: 2.5em; }
        table { border-collapse: collapse; width: 100%; margin: 1em 0; }
        th, td { border: 1px solid #eaecef; padding: 8px 12px; }
        th { background: #f0f6ff; color: #2d6cdf; }
        tr:nth-child(even) { background: #f6f8fa; }
        .risk-critical { background: #ffb3b3 !important; }
        .risk-high { background: #ffd699 !important; }
        .risk-medium { background: #fff7b3 !important; }
        .risk-low { background: #d6f5d6 !important; }
        .risk-info { background: #e6f0fa !important; }
        .badge { display: inline-block; padding: 2px 10px; border-radius: 8px; font-size: 0.9em; color: #fff; }
        .badge-critical { background: #e53935; }
        .badge-high { background: #fb8c00; }
        .badge-medium { background: #fbc02d; }
        .badge-low { background: #43a047; }
        .badge-info { background: #1976d2; }
        .json-block { background: #f6f8fa; border-radius: 8px; padding: 12px; font-family: 'Fira Mono', monospace; font-size: 0.95em; overflow-x: auto; }
        .timestamp { color: #888; font-size: 0.95em; }
        .summary-box { background: #f0f6ff; border-left: 5px solid #2d6cdf; padding: 16px 24px; border-radius: 8px; margin: 1em 0; }
        </style>
        """

        def risk_class(risk):
            return {
                "critical": "risk-critical",
                "high": "risk-high",
                "medium": "risk-medium",
                "low": "risk-low",
                "info": "risk-info"
            }.get(str(risk).lower(), "")

        def badge(risk):
            return f'<span class="badge badge-{risk.lower()}">{risk.capitalize()}</span>'

        def section_header(title, icon=""):
            return f'<h2>{icon} {title}</h2>'

        def render_table(headers, rows, row_classes=None):
            html = "<table><tr>"
            for h in headers:
                html += f"<th>{h}</th>"
            html += "</tr>"
            for i, row in enumerate(rows):
                cls = row_classes[i] if row_classes and i < len(row_classes) else ""
                html += f'<tr class="{cls}">'
                for cell in row:
                    html += f"<td>{cell}</td>"
                html += "</tr>"
            html += "</table>"
            return html

        def render_json_block(obj):
            import json
            return f'<div class="json-block">{json.dumps(obj, indent=2, ensure_ascii=False)}</div>'

        html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>APTES Full Security Report - {self.target}</title>
{css}
</head>
<body>
<div class="container">
<h1>APTES Full Security Report</h1>
<div class="timestamp">Target: <b>{self.target}</b> &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
"""

        # Recon Section
        recon = self.results.get("recon", {})
        html += '<div class="section">'
        html += section_header("Reconnaissance Phase", "🔍")
        if recon:
            passive = recon.get("passive", {})
            html += '<div class="summary-box"><b>Passive Reconnaissance</b><br>'
            html += f"DNS Records: {', '.join(passive.get('dns', {}).get('A', [])) or 'None'}<br>"
            html += f"Subdomains: {', '.join(passive.get('subdomains', [])) or 'None'}<br>"
            ssl = passive.get("ssl_info", {})
            if ssl:
                html += f"SSL Issuer: {ssl.get('issuer', 'N/A')}, Expiry: {ssl.get('expiry', 'N/A')}<br>"
            html += '</div>'
            active = recon.get("active", {})
            if "ports" in active:
                html += section_header("Port Scan Results", "🖧")
                port_rows, port_classes = [], []
                for host, host_data in active["ports"].items():
                    for proto, proto_data in host_data.items():
                        for port, port_data in proto_data.items():
                            port_rows.append([
                                host, proto, port, port_data.get("state", ""), port_data.get("service", "")
                            ])
                            port_classes.append("")
                if port_rows:
                    html += render_table(["Host", "Protocol", "Port", "State", "Service"], port_rows, port_classes)
            if "web" in active:
                html += section_header("Web Scan Results", "🌐")
                web = active["web"]
                # Technologies
                if "technologies" in web and web["technologies"]:
                    html += "<b>Technologies Detected:</b> " + ", ".join(web["technologies"]) + "<br>"
                # Headers
                if "headers" in web:
                    html += "<b>HTTP Headers:</b>"
                    for url, headers in web["headers"].items():
                        html += f"<br><b>{url}</b>:<ul style='margin-top:0'>"
                        for k, v in headers.items():
                            html += f"<li><b>{k}:</b> {v}</li>"
                        html += "</ul>"
                # Interesting files
                if "interesting_files" in web and web["interesting_files"]:
                    html += "<b>Interesting Files:</b><ul>"
                    for f in web["interesting_files"]:
                        html += f"<li><b>{f['url']}</b>: {f['content'][:100]}</li>"
                    html += "</ul>"
            webcrawler = recon.get("webcrawler", {})
            if webcrawler:
                html += section_header("Web Crawler Results", "🕷️")
                html += f"<b>Crawled URLs:</b> {webcrawler.get('total_urls_crawled', 0)}<br>"
                html += f"<b>Forms Found:</b> {webcrawler.get('total_forms_found', 0)}<br>"
                if "findings" in webcrawler and webcrawler["findings"]:
                    html += "<b>Findings:</b><ul>"
                    for finding in webcrawler["findings"]:
                        html += f"<li>{finding.get('finding', '')} ({finding.get('category', '')}) - {finding.get('description', '')}</li>"
                    html += "</ul>"
        else:
            html += "<i>No reconnaissance data available.</i>"
        html += '</div>'

        # Pre-Exploit Section
        pre = self.results.get("preexploit", {})
        html += '<div class="section">'
        html += section_header("Pre-Exploitation Phase", "🛡️")
        if pre:
            vulnval = pre.get("vulnerability_validation", {})
            vulns = vulnval.get("vulnerabilities", [])
            if vulns:
                html += section_header("Validated Vulnerabilities", "⚠️")
                vuln_rows, vuln_classes = [], []
                for v in vulns:
                    risk = v.get("risk_level", "info")
                    vuln_rows.append([
                        v.get("vulnerability", "Unknown"),
                        v.get("host", ""),
                        v.get("port", ""),
                        v.get("service", ""),
                        badge(risk),
                        v.get("cve", ""),
                        v.get("details", "")[:80]
                    ])
                    vuln_classes.append(risk_class(risk))
                html += render_table(
                    ["Name", "Host", "Port", "Service", "Risk", "CVE", "Details"], vuln_rows, vuln_classes
                )
            webapp = pre.get("webapp_analysis", {})
            findings = webapp.get("findings", [])
            if findings:
                html += section_header("Web Application Findings", "🔎")
                find_rows, find_classes = [], []
                for f in findings:
                    risk = f.get("risk_level", "info")
                    find_rows.append([
                        f.get("finding", ""),
                        f.get("url", ""),
                        f.get("category", ""),
                        badge(risk),
                        f.get("description", "")[:80],
                        f.get("recommendation", "")[:60]
                    ])
                    find_classes.append(risk_class(risk))
                html += render_table(
                    ["Finding", "URL", "Category", "Risk", "Description", "Recommendation"], find_rows, find_classes
                )
            attacks = pre.get("attack_vectors", [])
            if attacks:
                html += section_header("Attack Vectors", "🎯")
                atk_rows, atk_classes = [], []
                for a in attacks:
                    risk = a.get("risk_level", "info")
                    atk_rows.append([
                        a.get("type", ""),
                        a.get("name", ""),
                        a.get("target", ""),
                        badge(risk),
                        a.get("description", "")[:80]
                    ])
                    atk_classes.append(risk_class(risk))
                html += render_table(
                    ["Type", "Name", "Target", "Risk", "Description"], atk_rows, atk_classes
                )
            owasp = pre.get("owasp_top10", {})
            if owasp:
                html += section_header("OWASP Top 10 Coverage", "🔟")
                html += "<table><tr><th>OWASP Top 10</th><th>Status</th><th>Details</th></tr>"
                for key, val in owasp.items():
                    status = val.get("status", "") if isinstance(val, dict) else ""
                    details = val.get("details", "") if isinstance(val, dict) else ""
                    if isinstance(details, list):
                        details = "<ul>" + "".join(f"<li>{d}</li>" for d in details) + "</ul>"
                    html += f"<tr><td><b>{key}</b></td><td>{status}</td><td>{details}</td></tr>"
                html += "</table>"
        else:
            html += "<i>No pre-exploitation data available.</i>"
        html += '</div>'

        # Exploit Section
        exploit = self.results.get("exploit", {})
        html += '<div class="section">'
        html += section_header("Exploitation Phase", "💥")
        # Show exploitation summary if present and not all zero
        summary = exploit.get("exploitation_summary", {})
        if summary and any(summary.values()):
            html += "<h3>Exploitation Summary</h3>"
            html += "<ul>"
            html += f"<li><b>Attempts:</b> {summary.get('attempts', 0)}</li>"
            html += f"<li><b>Successful:</b> {summary.get('successful', 0)}</li>"
            html += f"<li><b>Failed:</b> {summary.get('failed', 0)}</li>"
            html += "</ul>"
        # Show exploits table if present
        exploits = exploit.get("exploits", [])
        if exploits:
            html += "<h3>Exploits Used</h3>"
            html += "<table><tr><th>Name</th><th>Target</th><th>Success</th><th>Details</th></tr>"
            for e in exploits:
                html += f"<tr><td>{e.get('name','')}</td><td>{e.get('target','')}</td><td>{'Yes' if e.get('success',False) else 'No'}</td><td>{e.get('details','')}</td></tr>"
            html += "</table>"
        # Show shells table if present
        shells = exploit.get("shells", [])
        if shells:
            html += "<h3>Obtained Shells</h3>"
            html += "<table><tr><th>Type</th><th>Target</th><th>Port</th><th>Privileges</th><th>Notes</th></tr>"
            for s in shells:
                html += f"<tr><td>{s.get('type','')}</td><td>{s.get('target','')}</td><td>{s.get('port','')}</td><td>{s.get('privileges','')}</td><td>{s.get('notes','')}</td></tr>"
            html += "</table>"
        # Show privilege escalation if present
        pe = exploit.get("privilege_escalation", [])
        if pe:
            html += "<h3>Privilege Escalation</h3>"
            for p in pe:
                html += f"<div><b>Technique:</b> {p.get('technique','')}<br>"
                html += f"<b>Target:</b> {p.get('target','')}<br>"
                html += f"<b>Initial User:</b> {p.get('initial_user','')}<br>"
                html += f"<b>Escalated User:</b> {p.get('escalated_user','')}<br>"
                html += f"<b>Details:</b> {p.get('details','')}</div><br>"
        # Always show LLM suggestions if present (even if no exploits/shells/pe)
        llm_suggestions = exploit.get("llm_suggestions")
        if llm_suggestions and (
            (isinstance(llm_suggestions, dict) and llm_suggestions.get("raw_response"))
            or (isinstance(llm_suggestions, str) and llm_suggestions.strip())
        ):
            html += self._render_llm_suggestions(llm_suggestions)
        elif not (summary or exploits or shells or pe):
            html += "<i>No exploitation data available.</i>"
        html += '</div>'

        # Post-Exploitation Section (optional, for completeness)
        post = self.results.get("postexploit", {})
        html += '<div class="section">'
        html += section_header("Post-Exploitation Phase", "🔓")
        if post:
            html += render_json_block(post)
        else:
            html += "<i>No post-exploitation data available.</i>"
        html += '</div>'

        html += f"""
        <div class="timestamp">Report generated by APTES on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
        </body>
        </html>
        """

        with open(filename, "w", encoding="utf-8") as f:
            f.write(html)
        return filename

def wait_and_generate_final_html_report(json_path, html_path, target, output_dir):
    """
    Wait for the final JSON file to appear, then generate the final HTML report.
    Args:
        json_path (str): Path to the final JSON file.
        html_path (str): Path to write the final HTML report.
        target (str): Target name.
        output_dir (str): Output directory for reports.
    """
    logger.info(f"Waiting for final JSON result: {json_path}")
    waited = 0
    max_wait = 600  # 10 minutes
    while not os.path.exists(json_path) and waited < max_wait:
        time.sleep(2)
        waited += 2
    if os.path.exists(json_path):
        logger.info(f"Final JSON found: {json_path}, generating HTML report: {html_path}")
        with open(json_path, "r", encoding="utf-8") as f:
            results = json.load(f)
        rg = ReportGenerator(results, target, output_dir)
        rg.generate_combined_html_report(filename=html_path)
        logger.info(f"Final HTML report generated: {html_path}")
    else:
        logger.warning(f"Timeout waiting for final JSON: {json_path}")

def start_final_report_subprocess(target, output_dir="reports"):
    """
    Start a background thread to wait for the final JSON and generate HTML.
    Call this after all phases are complete.
    """
    today = date.today().strftime("%Y%m%d")
    target_safe = target.replace(".", "_").replace(":", "_").replace("/", "_")
    base = f"{output_dir}/{target_safe}_recon_preexploit_exploit_{today}"
    json_path = f"{base}.json"
    html_path = f"{base}.html"
    t = threading.Thread(
        target=wait_and_generate_final_html_report,
        args=(json_path, html_path, target, output_dir),
        daemon=True
    )
    t.start()
    logger.info("Started background thread for final HTML report generation.")
