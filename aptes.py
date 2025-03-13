#!/usr/bin/env python3
"""
APTES - Advanced Penetration Testing and Exploitation Suite
----------------------------------------------------------
A comprehensive security assessment framework for reconnaissance,
pre-exploitation, exploitation, and post-exploitation phases.

Usage: python aptes.py [target] [options]
"""

import os
import re
import sys
import json
import time
import base64
import random
import string
import logging
import argparse
import subprocess
import socket
import urllib.parse
import hashlib
import urllib3
from datetime import datetime
from pathlib import Path

# Optional imports - will be checked before use
try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import concurrent.futures
    CONCURRENT_AVAILABLE = True
except ImportError:
    CONCURRENT_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('aptes')

class APTESFramework:
    """Main APTES Framework Class"""
    
    def __init__(self, target=None, output_dir="reports", threads=3, verbosity=1, verify_ssl=True):
        """Initialize the APTES framework"""
        self.target = target
        self.output_dir = output_dir
        self.threads = threads
        self.verbosity = verbosity
        self.verify_ssl = verify_ssl
        
        # Configure logging based on verbosity
        if verbosity >= 2:
            logging.getLogger().setLevel(logging.DEBUG)
        elif verbosity == 0:
            logging.getLogger().setLevel(logging.WARNING)
        
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"Created output directory: {output_dir}")
        
        # Initialize results dictionary properly
        self.results = {
            "target": target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "recon": {},
            "preexploit": {},
            "exploit": {},
            "postexploit": {}
        }
        
        # Initialize phase controllers with proper reference to framework
        self.recon = ReconnaissancePhase(self)
        self.preexploit = PreExploitationPhase(self)
        self.exploit = ExploitationPhase(self)
        self.postexploit = PostExploitationPhase(self)
        
        # Current phase tracking
        self.current_phase = None
    
    def run_phase(self, phase_name, **kwargs):
        """Run a specific phase of the framework"""
        if phase_name not in ["recon", "preexploit", "exploit", "postexploit"]:
            logger.error(f"Unknown phase: {phase_name}")
            return False
        
        self.current_phase = phase_name
        logger.info(f"Starting {phase_name} phase for {self.target}")
        
        start_time = time.time()
        phase_controller = getattr(self, phase_name)
        
        # Run the phase
        try:
            # Run the phase and get results
            result = phase_controller.run(**kwargs)
            
            # Store results in the framework's results dictionary
            self.results[phase_name] = result
            
            # Calculate duration
            end_time = time.time()
            duration = end_time - start_time
            self.results[phase_name]["duration"] = duration
            
            logger.info(f"{phase_name.capitalize()} phase completed in {duration:.2f} seconds")
            return True
        except Exception as e:
            logger.error(f"Error in {phase_name} phase: {str(e)}")
            if self.verbosity >= 2:
                import traceback
                traceback.print_exc()
            return False
    
    def save_results(self, filename=None):
        """Save all results to a JSON file"""
        if filename is None:
            # Create a default filename based on target and timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target_safe = self.target.replace(".", "_").replace(":", "_")
            filename = f"{self.output_dir}/{target_safe}_{timestamp}.json"
        
        with open(filename, "w") as f:
            json.dump(self.results, f, indent=4)
        
        logger.info(f"Results saved to {filename}")
        return filename
    
    def load_results(self, filename):
        """Load results from a JSON file"""
        try:
            with open(filename, "r") as f:
                data = json.load(f)
            
            # Update results with loaded data
            self.results.update(data)
            
            # Set target if not already set
            if not self.target and "target" in data:
                self.target = data["target"]
            
            logger.info(f"Loaded results from {filename}")
            return True
        except Exception as e:
            logger.error(f"Error loading results from {filename}: {str(e)}")
            return False
    
    def print_banner(self):
        """Print the APTES banner"""
        banner = """
        ╔═══════════════════════════════════════════════════════╗
        ║             APTES - Advanced Penetration               ║
        ║             Testing and Exploitation Suite             ║
        ╚═══════════════════════════════════════════════════════╝
        """
        print(banner)
        print(f"  Target: {self.target}")
        print(f"  Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("  " + "="*53)

class ReconnaissancePhase:
    """Reconnaissance Phase Controller"""
    
    def __init__(self, framework):
        """Initialize the Reconnaissance phase controller"""
        self.framework = framework
        self.target = framework.target
        self.output_dir = framework.output_dir
        self.threads = framework.threads
        self.results = {
            "target": self.target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "passive": {},
            "active": {}
        }
    
    def run(self, passive_only=False, skip_web=False):
        """Run reconnaissance phase"""
        logger.info(f"Running reconnaissance on {self.target}")
        
        # Perform passive reconnaissance
        self.passive_recon()
        
        # Perform active reconnaissance if not passive only
        if not passive_only:
            self.active_recon(skip_web=skip_web)
        
        return self.results
    
    def passive_recon(self):
        """Perform passive reconnaissance"""
        logger.info("Starting passive reconnaissance")
        
        # Check if we have concurrent futures for parallel execution
        if CONCURRENT_AVAILABLE:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
                dns_future = executor.submit(self.dns_lookup)
                subdomain_future = executor.submit(self.subdomain_enum)
                ssl_future = executor.submit(self.ssl_info)
                
                self.results["passive"]["dns"] = dns_future.result()
                self.results["passive"]["subdomains"] = subdomain_future.result()
                self.results["passive"]["ssl_info"] = ssl_future.result()
        else:
            # Sequential execution if concurrent not available
            self.results["passive"]["dns"] = self.dns_lookup()
            self.results["passive"]["subdomains"] = self.subdomain_enum()
            self.results["passive"]["ssl_info"] = self.ssl_info()
        
        logger.info("Passive reconnaissance completed")
        return self.results["passive"]
    
    def active_recon(self, skip_web=False):
        """Perform active reconnaissance"""
        logger.info("Starting active reconnaissance")
        
        # Check if we have concurrent futures for parallel execution
        if CONCURRENT_AVAILABLE:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
                port_scan_future = executor.submit(self.port_scan)
                service_enum_future = executor.submit(self.service_enum)
                vuln_scan_future = executor.submit(self.vuln_scan)
                
                self.results["active"]["ports"] = port_scan_future.result()
                self.results["active"]["services"] = service_enum_future.result()
                self.results["active"]["vulnerabilities"] = vuln_scan_future.result()
        else:
            # Sequential execution if concurrent not available
            self.results["active"]["ports"] = self.port_scan()
            self.results["active"]["services"] = self.service_enum()
            self.results["active"]["vulnerabilities"] = self.vuln_scan()
        
        # Perform web scanning if not skipped
        if not skip_web:
            self.results["active"]["web"] = self.web_scan()
        
        logger.info("Active reconnaissance completed")
        return self.results["active"]
    
    def dns_lookup(self):
        """Perform DNS lookups"""
        logger.info(f"Performing DNS lookups for {self.target}")
        record_types = ["A", "AAAA", "MX", "NS", "TXT", "SOA"]
        results = {}
        
        for record_type in record_types:
            try:
                cmd = f"dig +short {self.target} {record_type}"
                output = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
                
                if output:
                    results[record_type] = output.split('\n')
                else:
                    results[record_type] = []
            except Exception as e:
                logger.debug(f"Error in DNS lookup for {record_type}: {str(e)}")
                results[record_type] = []
                
        return results
    
    def subdomain_enum(self):
        """Enumerate subdomains"""
        logger.info(f"Enumerating subdomains for {self.target}")
        subdomains = set()
        
        # Try to use crt.sh for subdomain enumeration
        if REQUESTS_AVAILABLE:
            try:
                response = requests.get(f"https://crt.sh/?q=%.{self.target}&output=json")
                if response.status_code == 200:
                    data = response.json()
                    for entry in data:
                        domain_name = entry['name_value']
                        subdomains.add(domain_name)
            except Exception as e:
                logger.debug(f"Error in crt.sh lookup: {str(e)}")
        
        # Try to use subfinder if available
        try:
            cmd = f"subfinder -d {self.target} -silent"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
            if output:
                for subdomain in output.split('\n'):
                    subdomains.add(subdomain.strip())
        except Exception as e:
            logger.debug(f"Error using subfinder: {str(e)}")
        
        # Try to use dnsrecon if available
        try:
            cmd = f"dnsrecon -d {self.target} -t brt"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
            if output:
                # Extract domains from dnsrecon output
                domain_matches = re.findall(r"([a-zA-Z0-9\-\.]+\.{0})".format(self.target), output)
                for domain in domain_matches:
                    subdomains.add(domain.strip())
        except Exception as e:
            logger.debug(f"Error using dnsrecon: {str(e)}")
            
        return list(subdomains)
    
    def ssl_info(self):
        """Gather SSL certificate information"""
        logger.info(f"Gathering SSL certificate information for {self.target}")
        try:
            cmd = f"echo | openssl s_client -servername {self.target} -connect {self.target}:443 2>/dev/null | openssl x509 -noout -text"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            result = {}
            issuer_match = re.search(r"Issuer:.*?=(.*?)[,\n]", output)
            if issuer_match:
                result["issuer"] = issuer_match.group(1).strip()
                
            expiry_match = re.search(r"Not After\s*:\s*(.*?)\n", output)
            if expiry_match:
                result["expiry"] = expiry_match.group(1).strip()
                
            subject_alt_match = re.search(r"Subject Alternative Name:.*?DNS:(.*?)(?:,|$)", output)
            if subject_alt_match:
                result["alt_names"] = subject_alt_match.group(1).strip()
                
            return result
        except Exception as e:
            logger.debug(f"Error gathering SSL info: {str(e)}")
            return {"error": str(e)}
    
    def port_scan(self):
        """Perform port scanning"""
        logger.info(f"Scanning ports on {self.target}")
        try:
            cmd = f"nmap -F -T4 {self.target} -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            return results
        except Exception as e:
            logger.error(f"Error in port scanning: {str(e)}")
            
            # Fallback to simple TCP port scan if nmap fails
            results = self._simple_port_scan()
            return results
    
    def _simple_port_scan(self):
        """Perform a simple TCP port scan without nmap"""
        logger.info("Falling back to simple TCP port scan")
        common_ports = [21, 22, 23, 25, 53, 80, 110, 111, 135, 139, 143, 443, 445, 993, 995, 1723, 3306, 3389, 5900, 8080]
        results = {self.target: {"tcp": {}}}
        
        for port in common_ports:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex((self.target, port))
                if result == 0:
                    service = self._guess_service(port)
                    results[self.target]["tcp"][str(port)] = {
                        "port": str(port),
                        "state": "open",
                        "service": service
                    }
                sock.close()
            except Exception:
                pass
                
        return results
    
    def _guess_service(self, port):
        """Guess service based on port number"""
        service_map = {
            21: "ftp",
            22: "ssh",
            23: "telnet",
            25: "smtp",
            53: "domain",
            80: "http",
            110: "pop3",
            111: "rpcbind",
            135: "msrpc",
            139: "netbios-ssn",
            143: "imap",
            443: "https",
            445: "microsoft-ds",
            993: "imaps",
            995: "pop3s",
            1723: "pptp",
            3306: "mysql",
            3389: "ms-wbt-server",
            5900: "vnc",
            8080: "http-proxy"
        }
        return service_map.get(port, "unknown")
    
    def service_enum(self):
        """Enumerate services on open ports"""
        logger.info(f"Enumerating services on {self.target}")
        try:
            # Resolve domain to IP if needed
            target_ip = self.target
            if not self._is_ip_address(self.target):
                try:
                    target_ip = socket.gethostbyname(self.target)
                    logger.debug(f"Resolved {self.target} to {target_ip}")
                except socket.gaierror:
                    pass
            
            cmd = f"nmap -sV -T4 --script=banner {target_ip} -p 21,22,25,80,443,8080,8443 -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            return results
        except Exception as e:
            logger.error(f"Error in service enumeration: {str(e)}")
            return {"error": str(e)}
    
    def vuln_scan(self):
        """Perform vulnerability scanning"""
        logger.info(f"Checking for vulnerabilities on {self.target}")
        try:
            # Resolve domain to IP if needed
            target_ip = self.target
            if not self._is_ip_address(self.target):
                try:
                    target_ip = socket.gethostbyname(self.target)
                    logger.debug(f"Resolved {self.target} to {target_ip}")
                except socket.gaierror:
                    pass
            
            cmd = f"nmap -sV --script=vuln {target_ip} -p 21,22,25,80,443,8080,8443 -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            
            # Extract vulnerabilities from results
            vulns = []
            for host_data in results.values():
                for proto_data in host_data.values():
                    for port_data in proto_data.values():
                        scripts = port_data.get("scripts", {})
                        for script_name, script_output in scripts.items():
                            if "VULNERABLE" in script_output:
                                vulns.append({
                                    "port": port_data.get("port", "unknown"),
                                    "service": port_data.get("service", "unknown"),
                                    "vulnerability": script_name,
                                    "details": script_output
                                })
            
            return {"vulnerabilities": vulns, "scan_data": results}
        except Exception as e:
            logger.error(f"Error in vulnerability scanning: {str(e)}")
            return {"error": str(e)}
    
    def web_scan(self):
        """Perform web scanning"""
        logger.info(f"Performing web scan on {self.target}")
        
        if not REQUESTS_AVAILABLE:
            logger.error("Requests library not available, skipping web scan")
            return {"error": "Requests library not available"}
        
        results = {
            "headers": {},
            "technologies": [],
            "interesting_files": []
        }
        
        for port in [80, 443, 8080, 8443]:
            protocol = "https" if port in [443, 8443] else "http"
            url = f"{protocol}://{self.target}:{port}"
            
            try:
                response = requests.get(url, timeout=10, verify=self.verify_ssl)
                
                results["headers"][f"{url}"] = dict(response.headers)
                
                if "X-Powered-By" in response.headers:
                    results["technologies"].append(response.headers["X-Powered-By"])
                if "Server" in response.headers:
                    results["technologies"].append(response.headers["Server"])
                
                # Check for robots.txt
                try:
                    robots = requests.get(f"{url}/robots.txt", timeout=5, verify=False)
                    if robots.status_code == 200:
                        results["interesting_files"].append({
                            "url": f"{url}/robots.txt",
                            "content": robots.text[:500] 
                        })
                except Exception:
                    pass
                
                # Check for sitemap.xml
                try:
                    sitemap = requests.get(f"{url}/sitemap.xml", timeout=5, verify=False)
                    if sitemap.status_code == 200:
                        results["interesting_files"].append({
                            "url": f"{url}/sitemap.xml",
                            "content": "Sitemap found"
                        })
                except Exception:
                    pass
                
            except requests.exceptions.RequestException:
                continue
                
        return results
    
    def _parse_nmap_xml(self, xml_data):
        """Parse nmap XML output to extract port and service information"""
        results = {}
        
        host_blocks = re.findall(r"<host.*?</host>", xml_data, re.DOTALL)
        
        for host_block in host_blocks:
            addr_match = re.search(r'<address addr="([^"]*)"', host_block)
            if addr_match:
                host = addr_match.group(1)
                results[host] = {"tcp": {}}
                
                port_blocks = re.findall(r"<port.*?</port>", host_block, re.DOTALL)
                for port_block in port_blocks:
                    port_match = re.search(r'portid="([^"]*)"', port_block)
                    proto_match = re.search(r'protocol="([^"]*)"', port_block)
                    state_match = re.search(r'<state state="([^"]*)"', port_block)
                    service_match = re.search(r'<service name="([^"]*)"', port_block)
                    product_match = re.search(r'product="([^"]*)"', port_block)
                    version_match = re.search(r'version="([^"]*)"', port_block)
                    
                    if port_match and proto_match and state_match:
                        port = port_match.group(1)
                        proto = proto_match.group(1)
                        
                        if proto not in results[host]:
                            results[host][proto] = {}
                            
                        results[host][proto][port] = {
                            "port": port,
                            "state": state_match.group(1),
                            "service": service_match.group(1) if service_match else "unknown"
                        }
                        
                        if product_match:
                            results[host][proto][port]["product"] = product_match.group(1)
                        if version_match:
                            results[host][proto][port]["version"] = version_match.group(1)
                        
                        script_blocks = re.findall(r"<script.*?</script>", port_block, re.DOTALL)
                        if script_blocks:
                            scripts = {}
                            for script_block in script_blocks:
                                script_id_match = re.search(r'id="([^"]*)"', script_block)
                                script_output_match = re.search(r'output="([^"]*)"', script_block)
                                if script_id_match and script_output_match:
                                    scripts[script_id_match.group(1)] = script_output_match.group(1)
                            
                            if scripts:
                                results[host][proto][port]["scripts"] = scripts
        
        return results
    
    def _is_ip_address(self, address):
        """Check if the given address is an IP address"""
        try:
            socket.inet_aton(address)
            return True
        except socket.error:
            return False

class PreExploitationPhase:
    """Pre-Exploitation Phase Controller"""
    
    def __init__(self, framework):
        """Initialize the Pre-Exploitation phase controller"""
        self.framework = framework
        self.target = framework.target
        self.output_dir = framework.output_dir
        self.threads = framework.threads
        self.recon_results = framework.results.get("recon", {})
        self.results = {
            "target": self.target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "vulnerability_validation": {},
            "webapp_analysis": {},
            "credential_testing": {},
            "payload_generation": {},
            "attack_vectors": []
        }
        
        # Risk colors for reports
        self.risk_colors = {
            "critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") if EXCEL_AVAILABLE else None,
            "high": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") if EXCEL_AVAILABLE else None,
            "medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") if EXCEL_AVAILABLE else None,
            "low": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") if EXCEL_AVAILABLE else None,
            "info": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") if EXCEL_AVAILABLE else None
        }
    
    def run(self, exploit_filter=None, skip_web=False, skip_creds=False, skip_payloads=False):
        """Run pre-exploitation phase"""
        logger.info(f"Starting pre-exploitation analysis for {self.target}")
        
        # Validate vulnerabilities
        self.validate_vulnerabilities()
        
        # Perform web application analysis if not skipped
        if not skip_web:
            self.analyze_web_applications()
        
        # Test default credentials if not skipped
        if not skip_creds:
            self.test_default_credentials()
        
        # Generate payloads if not skipped
        if not skip_payloads:
            self.generate_payloads()
        
        # Identify attack vectors
        self.identify_attack_vectors(exploit_filter)
        
        logger.info(f"Pre-exploitation analysis completed for {self.target}")
        return self.results
    
    def validate_vulnerabilities(self):
        """Validate vulnerabilities from reconnaissance phase"""
        logger.info("Validating vulnerabilities")
        
        validated_vulns = []
        
        # Extract vulnerabilities from recon results
        if self.recon_results and "active" in self.recon_results and "vulnerabilities" in self.recon_results["active"]:
            vulns = self.recon_results["active"]["vulnerabilities"].get("vulnerabilities", [])
            
            # Process each vulnerability
            for vuln in vulns:
                validated_vuln = self._validate_vulnerability(vuln)
                if validated_vuln:
                    validated_vulns.append(validated_vuln)
        
        # Add manual vulnerability checks
        if self.recon_results and "active" in self.recon_results and "ports" in self.recon_results["active"]:
            for host, host_data in self.recon_results["active"]["ports"].items():
                for proto, proto_data in host_data.items():
                    for port, port_data in proto_data.items():
                        service = port_data.get("service", "").lower()
                        product = port_data.get("product", "").lower()
                        version = port_data.get("version", "")
                        
                        # Check for potential vulnerabilities
                        manual_vulns = self._check_known_vulnerabilities(host, port, service, product, version)
                        validated_vulns.extend(manual_vulns)
        
        # Group vulnerabilities by host and service
        grouped_vulns = {}
        for vuln in validated_vulns:
            host = vuln.get("host", self.target)
            service = vuln.get("service", "unknown")
            
            if host not in grouped_vulns:
                grouped_vulns[host] = {}
            
            if service not in grouped_vulns[host]:
                grouped_vulns[host][service] = []
            
            grouped_vulns[host][service].append(vuln)
        
        self.results["vulnerability_validation"] = {
            "vulnerabilities": validated_vulns,
            "grouped": grouped_vulns,
            "total_count": len(validated_vulns)
        }
        
        logger.info(f"Validated {len(validated_vulns)} vulnerabilities")
        return validated_vulns
    
    def _validate_vulnerability(self, vuln):
        """Validate and enhance a single vulnerability"""
        port = vuln.get("port", "unknown")
        service = vuln.get("service", "unknown")
        vuln_name = vuln.get("vulnerability", "unknown")
        
        logger.debug(f"Validating {vuln_name} on {self.target}:{port} ({service})")
        
        # Enhance the vulnerability data
        enhanced_vuln = vuln.copy()
        enhanced_vuln["host"] = self.target
        enhanced_vuln["validated"] = False
        enhanced_vuln["validation_method"] = "manual"
        enhanced_vuln["risk_level"] = "medium"  # Default
        
        # Add CVE details if present in the vulnerability name
        cve_match = re.search(r'(CVE-\d{4}-\d{4,7})', vuln_name)
        if cve_match:
            enhanced_vuln["cve"] = cve_match.group(1)
            
            # Try to get CVE details
            try:
                if REQUESTS_AVAILABLE:
                    cve_id = enhanced_vuln["cve"]
                    response = requests.get(f"https://services.nvd.nist.gov/rest/json/cve/1.0/{cve_id}")
                    if response.status_code == 200:
                        data = response.json()
                        if "result" in data and "CVE_Items" in data["result"] and data["result"]["CVE_Items"]:
                            cve_data = data["result"]["CVE_Items"][0]
                            
                            # Extract CVSS score and severity
                            if "impact" in cve_data:
                                if "baseMetricV3" in cve_data["impact"]:
                                    enhanced_vuln["cvss_v3"] = cve_data["impact"]["baseMetricV3"]["cvssV3"]["baseScore"]
                                    enhanced_vuln["risk_level"] = cve_data["impact"]["baseMetricV3"]["cvssV3"]["baseSeverity"].lower()
                                elif "baseMetricV2" in cve_data["impact"]:
                                    enhanced_vuln["cvss_v2"] = cve_data["impact"]["baseMetricV2"]["cvssV2"]["baseScore"]
                                    
                                    # Map CVSS V2 score to risk level
                                    score = float(enhanced_vuln["cvss_v2"])
                                    if score >= 9.0:
                                        enhanced_vuln["risk_level"] = "critical"
                                    elif score >= 7.0:
                                        enhanced_vuln["risk_level"] = "high"
                                    elif score >= 4.0:
                                        enhanced_vuln["risk_level"] = "medium"
                                    else:
                                        enhanced_vuln["risk_level"] = "low"
            except Exception as e:
                logger.debug(f"Error fetching CVE details: {str(e)}")
        
        # Attempt validation based on service type
        if "http" in service or port in ["80", "443", "8080", "8443"]:
            validated = self._validate_web_vulnerability(enhanced_vuln)
            enhanced_vuln["validated"] = validated
            enhanced_vuln["validation_method"] = "web_request"
        elif "ssh" in service:
            enhanced_vuln["validation_method"] = "ssh_check"
        elif "ftp" in service:
            enhanced_vuln["validation_method"] = "ftp_check"
        
        return enhanced_vuln
    
    def _validate_web_vulnerability(self, vuln):
        """Validate a web vulnerability"""
        port = vuln.get("port", "80")
        protocol = "https" if port in ["443", "8443"] else "http"
        url = f"{protocol}://{self.target}:{port}"
        
        vuln_name = vuln.get("vulnerability", "").lower()
        
        # Validation is limited without actual exploitation
        # In a real implementation, would perform more detailed checks
        
        return False  # Default to not validated
    
    def _check_known_vulnerabilities(self, host, port, service, product, version):
        """Check for known vulnerabilities in services"""
        vulnerabilities = []
        
        # Database of known vulnerabilities
        vuln_db = {
            "apache": {
                "2.4.49": {
                    "cve": "CVE-2021-41773",
                    "name": "Apache HTTP Server Path Traversal",
                    "risk_level": "critical"
                },
                "2.4.50": {
                    "cve": "CVE-2021-42013",
                    "name": "Apache HTTP Server Path Traversal",
                    "risk_level": "critical"
                }
            },
            "nginx": {
                "1.3.9": {
                    "cve": "CVE-2013-4547",
                    "name": "NGINX HTTP URI Processing Security Bypass",
                    "risk_level": "high"
                }
            },
            "openssh": {
                "7.7": {
                    "cve": "CVE-2018-15473",
                    "name": "OpenSSH Username Enumeration",
                    "risk_level": "medium"
                }
            }
        }
        
        # Generic vulnerabilities by service type
        generic_vulns = {
            "http": [
                {
                    "name": "Potential SQL Injection",
                    "risk_level": "high",
                    "description": "Web applications may be vulnerable to SQL injection if user input is not properly sanitized."
                },
                {
                    "name": "Potential Cross-Site Scripting (XSS)",
                    "risk_level": "medium",
                    "description": "Web applications may be vulnerable to XSS if user input is not properly sanitized."
                }
            ],
            "ftp": [
                {
                    "name": "Anonymous FTP Access",
                    "risk_level": "medium",
                    "description": "FTP server may allow anonymous access, exposing files to unauthorized users."
                }
            ],
            "ssh": [
                {
                    "name": "Weak SSH Configurations",
                    "risk_level": "medium",
                    "description": "SSH server may have weak ciphers or authentication methods enabled."
                }
            ]
        }
        
        # Check product-specific vulnerabilities
        for db_product, versions in vuln_db.items():
            if db_product in product.lower():
                for vuln_version, vuln_info in versions.items():
                    if version.startswith(vuln_version):
                        vulnerabilities.append({
                            "host": host,
                            "port": port,
                            "service": service,
                            "product": product,
                            "version": version,
                            "vulnerability": vuln_info["name"],
                            "cve": vuln_info["cve"],
                            "risk_level": vuln_info["risk_level"],
                            "validated": False,
                            "validation_method": "version_check"
                        })
        
        # Add generic service vulnerabilities
        for db_service, vulns in generic_vulns.items():
            if db_service in service.lower():
                for vuln in vulns:
                    vulnerabilities.append({
                        "host": host,
                        "port": port,
                        "service": service,
                        "vulnerability": vuln["name"],
                        "risk_level": vuln["risk_level"],
                        "description": vuln["description"],
                        "validated": False,
                        "validation_method": "service_check"
                    })
        
        return vulnerabilities
    
    def analyze_web_applications(self):
        """Analyze web applications for vulnerabilities"""
        logger.info("Analyzing web applications")
        
        web_services = []
        
        # Extract web services from recon results
        if self.recon_results and "active" in self.recon_results and "ports" in self.recon_results["active"]:
            for host, host_data in self.recon_results["active"]["ports"].items():
                for proto, proto_data in host_data.items():
                    for port, port_data in proto_data.items():
                        service = port_data.get("service", "").lower()
                        if "http" in service or port in ["80", "443", "8080", "8443"]:
                            protocol = "https" if port in ["443", "8443"] or "https" in service else "http"
                            web_services.append({
                                "host": host,
                                "port": port,
                                "url": f"{protocol}://{host}:{port}"
                            })
        
        # If no web services found in recon, try common ports
        if not web_services:
            for port in [80, 443, 8080, 8443]:
                protocol = "https" if port in [443, 8443] else "http"
                web_services.append({
                    "host": self.target,
                    "port": str(port),
                    "url": f"{protocol}://{self.target}:{port}"
                })
        
        # Analyze each web service
        web_results = {
            "services": [],
            "findings": []
        }
        
        if CONCURRENT_AVAILABLE:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
                future_to_service = {executor.submit(self._analyze_web_service, service): service for service in web_services}
                
                for future in concurrent.futures.as_completed(future_to_service):
                    service = future_to_service[future]
                    try:
                        result = future.result()
                        if result:
                            web_results["services"].append(result["service"])
                            web_results["findings"].extend(result["findings"])
                    except Exception as e:
                        logger.error(f"Error analyzing web service {service.get('url')}: {str(e)}")
        else:
            # Sequential analysis if concurrent not available
            for service in web_services:
                try:
                    result = self._analyze_web_service(service)
                    if result:
                        web_results["services"].append(result["service"])
                        web_results["findings"].extend(result["findings"])
                except Exception as e:
                    logger.error(f"Error analyzing web service {service.get('url')}: {str(e)}")
        
        # Group findings by category
        grouped_findings = {}
        for finding in web_results["findings"]:
            category = finding.get("category", "Other")
            if category not in grouped_findings:
                grouped_findings[category] = []
            grouped_findings[category].append(finding)
        
        web_results["grouped_findings"] = grouped_findings
        web_results["total_findings"] = len(web_results["findings"])
        
        self.results["webapp_analysis"] = web_results
        logger.info(f"Web application analysis complete: {len(web_results['findings'])} findings")
        return web_results
    
    def _analyze_web_service(self, service):
        """Analyze a single web service"""
        if not REQUESTS_AVAILABLE:
            logger.error("Requests library not available, skipping web analysis")
            return None
        
        url = service.get("url")
        host = service.get("host")
        port = service.get("port")
        
        logger.debug(f"Analyzing web application at {url}")
        
        findings = []
        service_info = {
            "url": url,
            "host": host,
            "port": port,
            "headers": {},
            "technologies": [],
            "server": "",
            "status_code": 0
        }
        
        try:
            # Request the main page
            response = requests.get(url, timeout=10, verify=False, allow_redirects=True)
            service_info["status_code"] = response.status_code
            service_info["headers"] = dict(response.headers)
            service_info["final_url"] = response.url
            
            # Extract server information
            if "Server" in response.headers:
                service_info["server"] = response.headers["Server"]
            
            # Check for missing security headers
            security_headers = [
                "Strict-Transport-Security",
                "Content-Security-Policy",
                "X-Content-Type-Options",
                "X-Frame-Options",
                "X-XSS-Protection"
            ]
            
            for header in security_headers:
                if header not in response.headers:
                    findings.append({
                        "host": host,
                        "port": port,
                        "url": url,
                        "category": "Missing Security Headers",
                        "finding": f"Missing {header} header",
                        "risk_level": "low",
                        "description": f"The {header} security header is missing from the response.",
                        "recommendation": f"Implement the {header} header to enhance security."
                    })
            
            # Check for directory listing
            common_dirs = ["backup", "admin", "wp-admin", "config", "test", "dev"]
            for directory in common_dirs:
                try:
                    dir_url = f"{url}/{directory}/"
                    dir_response = requests.get(dir_url, timeout=5, verify=False)
                    
                    if dir_response.status_code == 200:
                        listing_patterns = [
                            "Index of /",
                            "Directory Listing For",
                            "<title>Index of",
                            "Parent Directory</a>"
                        ]
                        
                        for pattern in listing_patterns:
                            if pattern in dir_response.text:
                                findings.append({
                                    "host": host,
                                    "port": port,
                                    "url": dir_url,
                                    "category": "Directory Listing",
                                    "finding": f"Directory listing enabled at {directory}/",
                                    "risk_level": "medium",
                                    "description": "The web server is configured to show directory listings, which can expose sensitive files.",
                                    "recommendation": "Disable directory listings in the web server configuration."
                                })
                                break
                except Exception:
                    pass
            
            # Test for common web vulnerabilities
            self._test_web_vulnerabilities(url, host, port, findings)
            
        except Exception as e:
            logger.debug(f"Error analyzing {url}: {str(e)}")
        
        return {
            "service": service_info,
            "findings": findings
        }
    
    def _test_web_vulnerabilities(self, url, host, port, findings):
        """Test for common web vulnerabilities"""
        if not REQUESTS_AVAILABLE:
            return
        
        # Test for SQL injection
        sql_paths = ["/login", "/search", "/product", "/user"]
        sql_payloads = ["'", "' OR '1'='1", "1' OR '1'='1"]
        
        for path in sql_paths:
            for payload in sql_payloads:
                try:
                    test_url = f"{url}{path}?id={payload}"
                    response = requests.get(test_url, timeout=5, verify=False)
                    
                    # Look for SQL error messages
                    sql_errors = [
                        "sql syntax",
                        "syntax error",
                        "mysql_fetch",
                        "unclosed quotation mark",
                        "ORA-",
                        "pg_query"
                    ]
                    
                    for error in sql_errors:
                        if error.lower() in response.text.lower():
                            findings.append({
                                "host": host,
                                "port": port,
                                "url": test_url,
                                "category": "SQL Injection",
                                "finding": f"Potential SQL injection vulnerability at {path}",
                                "risk_level": "high",
                                "description": "The application may be vulnerable to SQL injection attacks. SQL error messages were detected in the response.",
                                "recommendation": "Implement proper input validation and parameterized queries."
                            })
                            break
                except Exception:
                    pass
        
        # Test for XSS
        xss_paths = ["/search", "/comment", "/feedback"]
        xss_payloads = ["<script>alert(1)</script>", "<img src=x onerror=alert(1)>"]
        
        for path in xss_paths:
            for payload in xss_payloads:
                try:
                    test_url = f"{url}{path}?q={urllib.parse.quote_plus(payload)}"
                    response = requests.get(test_url, timeout=5, verify=False)
                    
                    # Check if payload is reflected
                    if payload in response.text:
                        findings.append({
                            "host": host,
                            "port": port,
                            "url": test_url,
                            "category": "Cross-Site Scripting (XSS)",
                            "finding": f"Potential XSS vulnerability at {path}",
                            "risk_level": "medium",
                            "description": "The application may be vulnerable to Cross-Site Scripting attacks. User input is being reflected in the response without proper sanitization.",
                            "recommendation": "Implement proper output encoding and content security policy."
                        })
                        break
                except Exception:
                    pass
    
    def test_default_credentials(self):
        """Test for default credentials"""
        logger.info("Testing for default credentials")
        
        cred_results = {
            "tested_services": [],
            "findings": []
        }
        
        # Extract services to test from recon results
        services_to_test = []
        if self.recon_results and "active" in self.recon_results and "ports" in self.recon_results["active"]:
            for host, host_data in self.recon_results["active"]["ports"].items():
                for proto, proto_data in host_data.items():
                    for port, port_data in proto_data.items():
                        service = port_data.get("service", "").lower()
                        
                        if service in ["http", "https", "ssh", "ftp", "telnet", "mysql", "mssql"]:
                            services_to_test.append({
                                "host": host,
                                "port": port,
                                "service": service,
                                "product": port_data.get("product", ""),
                                "version": port_data.get("version", "")
                            })
        
        # If no services found in recon, test common services
        if not services_to_test:
            for service in ["http", "https", "ssh", "ftp"]:
                port = {"http": "80", "https": "443", "ssh": "22", "ftp": "21"}.get(service)
                services_to_test.append({
                    "host": self.target,
                    "port": port,
                    "service": service
                })
        
        # Test each service
        if CONCURRENT_AVAILABLE:
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.threads) as executor:
                future_to_service = {executor.submit(self._test_service_credentials, service): service for service in services_to_test}
                
                for future in concurrent.futures.as_completed(future_to_service):
                    service = future_to_service[future]
                    try:
                        result = future.result()
                        if result:
                            cred_results["tested_services"].append({
                                "host": service["host"],
                                "port": service["port"],
                                "service": service["service"]
                            })
                            
                            if result.get("findings"):
                                cred_results["findings"].extend(result["findings"])
                    except Exception as e:
                        logger.error(f"Error testing credentials for {service['service']} on {service['host']}:{service['port']}: {str(e)}")
        else:
            # Sequential testing if concurrent not available
            for service in services_to_test:
                try:
                    result = self._test_service_credentials(service)
                    if result:
                        cred_results["tested_services"].append({
                            "host": service["host"],
                            "port": service["port"],
                            "service": service["service"]
                        })
                        
                        if result.get("findings"):
                            cred_results["findings"].extend(result["findings"])
                except Exception as e:
                    logger.error(f"Error testing credentials for {service['service']} on {service['host']}:{service['port']}: {str(e)}")
        
        cred_results["total_tested"] = len(cred_results["tested_services"])
        cred_results["total_findings"] = len(cred_results["findings"])
        
        self.results["credential_testing"] = cred_results
        logger.info(f"Credential testing complete: {len(cred_results['findings'])} findings from {len(cred_results['tested_services'])} services")
        return cred_results
    
    def _test_service_credentials(self, service):
        """Test credentials for a specific service"""
        host = service["host"]
        port = service["port"]
        service_name = service["service"]
        
        logger.debug(f"Testing credentials for {service_name} on {host}:{port}")
        
        findings = []
        
        # Service-specific credential testing
        if service_name in ["http", "https"]:
            findings.extend(self._test_web_credentials(host, port, service_name))
        elif service_name == "ftp":
            findings.extend(self._test_ftp_credentials(host, port))
        elif service_name == "ssh":
            findings.extend(self._test_ssh_credentials(host, port))
        
        return {
            "service": {
                "host": host,
                "port": port,
                "name": service_name
            },
            "findings": findings
        }
    
    def _test_web_credentials(self, host, port, service):
        """Test credentials for web services"""
        if not REQUESTS_AVAILABLE:
            return []
        
        findings = []
        protocol = "https" if service == "https" or port in ["443", "8443"] else "http"
        
        # Common admin paths
        admin_paths = [
            "/admin",
            "/administrator",
            "/login",
            "/wp-admin",
            "/admin-panel"
        ]
        
        # Default credentials to test
        default_creds = [
            {"username": "admin", "password": "admin"},
            {"username": "admin", "password": "password"},
            {"username": "administrator", "password": "administrator"},
            {"username": "root", "password": "root"}
        ]
        
        # Check each admin path
        for path in admin_paths:
            url = f"{protocol}://{host}:{port}{path}"
            
            try:
                # First check if the page exists
                response = requests.get(url, timeout=5, verify=self.verify_ssl)
                
                # Look for login forms
                if response.status_code == 200 and ("login" in response.text.lower() or "password" in response.text.lower()):
                    logger.debug(f"Found potential login page at {url}")
                    
                    # In a real implementation, would attempt logins
                    # Here we'll just report it as a finding to check manually
                    findings.append({
                        "host": host,
                        "port": port,
                        "service": service,
                        "finding": f"Potential login page found at {url}",
                        "category": "Default Credentials",
                        "risk_level": "info",
                        "description": "A login page was identified which should be checked for default or weak credentials.",
                        "credentials_to_try": [f"{cred['username']}:{cred['password']}" for cred in default_creds[:3]],
                        "recommendation": "Try default credentials and ensure strong password policies are enforced."
                    })
            except Exception:
                pass
        
        return findings
    
    def _test_ftp_credentials(self, host, port):
        """Test credentials for FTP services"""
        findings = []
        
        # Try anonymous access
        try:
            cmd = f"timeout 10 ftp -n {host} {port} <<EOF\nuser anonymous anonymous\nls\nquit\nEOF"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            if "230" in output:  # 230 = Login successful
                findings.append({
                    "host": host,
                    "port": port,
                    "service": "ftp",
                    "finding": "Anonymous FTP access allowed",
                    "category": "Default Credentials",
                    "risk_level": "medium",
                    "description": "The FTP server allows anonymous access, which could expose sensitive files.",
                    "credentials": "anonymous:anonymous",
                    "recommendation": "Disable anonymous FTP access unless specifically required."
                })
        except Exception:
            pass
        
        return findings
    
    def _test_ssh_credentials(self, host, port):
        """Test credentials for SSH services"""
        # In a real implementation, would attempt common credentials
        # Here we'll just return a suggestion to check manually
        
        findings = [{
            "host": host,
            "port": port,
            "service": "ssh",
            "finding": "SSH service should be checked for default credentials",
            "category": "Default Credentials",
            "risk_level": "info",
            "description": "The SSH service should be manually checked for default or weak credentials.",
            "credentials_to_try": ["root:root", "admin:admin", "user:password"],
            "recommendation": "Ensure strong password policies and consider using key-based authentication instead of passwords."
        }]
        
        return findings
    
    def generate_payloads(self):
        """Generate payloads for identified vulnerabilities"""
        logger.info("Generating payloads for identified vulnerabilities")
        
        payload_results = {
            "payloads": [],
            "payload_types": set()
        }
        
        # Generate payloads for validated vulnerabilities
        if "vulnerability_validation" in self.results and "vulnerabilities" in self.results["vulnerability_validation"]:
            vulns = self.results["vulnerability_validation"]["vulnerabilities"]
            
            for vuln in vulns:
                if vuln.get("risk_level") in ["critical", "high"]:
                    payload = self._generate_payload_for_vulnerability(vuln)
                    if payload:
                        payload_results["payloads"].append(payload)
                        payload_results["payload_types"].add(payload["type"])
        
        # Generate payloads based on web findings
        if "webapp_analysis" in self.results and "findings" in self.results["webapp_analysis"]:
            for finding in self.results["webapp_analysis"]["findings"]:
                if finding.get("risk_level") in ["critical", "high", "medium"]:
                    category = finding.get("category")
                    
                    if category == "SQL Injection":
                        payload = self._generate_sql_injection_payload(finding)
                        if payload:
                            payload_results["payloads"].append(payload)
                            payload_results["payload_types"].add(payload["type"])
                    
                    elif category == "Cross-Site Scripting (XSS)":
                        payload = self._generate_xss_payload(finding)
                        if payload:
                            payload_results["payloads"].append(payload)
                            payload_results["payload_types"].add(payload["type"])
        
        # Set payload counts and convert set to list
        payload_results["total_payloads"] = len(payload_results["payloads"])
        payload_results["payload_types"] = list(payload_results["payload_types"])
        
        self.results["payload_generation"] = payload_results
        logger.info(f"Generated {len(payload_results['payloads'])} payloads")
        return payload_results
    
    def _generate_payload_for_vulnerability(self, vuln):
        """Generate payload for a specific vulnerability"""
        vuln_name = vuln.get("vulnerability", "").lower()
        
        # Map vulnerabilities to generators
        if "sql injection" in vuln_name:
            return self._generate_sql_injection_payload(vuln)
        elif "xss" in vuln_name or "cross-site scripting" in vuln_name:
            return self._generate_xss_payload(vuln)
        elif "lfi" in vuln_name or "file inclusion" in vuln_name:
            return self._generate_lfi_payload(vuln)
        elif "rce" in vuln_name or "code execution" in vuln_name:
            return self._generate_rce_payload(vuln)
        
        return None
    
    def _generate_sql_injection_payload(self, finding):
        """Generate SQL injection payloads"""
        return {
            "type": "sqli",
            "name": "SQL Injection Payload",
            "target": finding.get("host", self.target),
            "port": finding.get("port", "80"),
            "vulnerability": "SQL Injection",
            "description": "Payloads to exploit SQL injection vulnerabilities",
            "payloads": [
                "' OR 1=1 --",
                "' OR '1'='1",
                "admin' --",
                "' UNION SELECT 1,2,3,4,5 --",
                "' UNION SELECT username,password,3,4,5 FROM users --"
            ],
            "usage": "These payloads should be inserted into input fields like login forms, search boxes, or URL parameters.",
            "risk_level": "high"
        }
    
    def _generate_xss_payload(self, finding):
        """Generate XSS payloads"""
        return {
            "type": "xss",
            "name": "Cross-Site Scripting Payload",
            "target": finding.get("host", self.target),
            "port": finding.get("port", "80"),
            "vulnerability": "Cross-Site Scripting (XSS)",
            "description": "Payloads to exploit XSS vulnerabilities",
            "payloads": [
                "<script>alert('XSS')</script>",
                "<img src=x onerror=alert('XSS')>",
                "<body onload=alert('XSS')>",
                "<svg/onload=alert('XSS')>",
                "<iframe src=\"javascript:alert('XSS')\">"
            ],
            "usage": "These payloads should be inserted into input fields, URL parameters, or form submissions.",
            "risk_level": "medium"
        }
    
    def _generate_lfi_payload(self, finding):
        """Generate LFI payloads"""
        return {
            "type": "lfi",
            "name": "Local File Inclusion Payload",
            "target": finding.get("host", self.target),
            "port": finding.get("port", "80"),
            "vulnerability": "Local File Inclusion",
            "description": "Payloads to exploit local file inclusion vulnerabilities",
            "payloads": [
                "../../../etc/passwd",
                "../../../etc/shadow",
                "../../../proc/self/environ",
                "../../../var/log/apache2/access.log",
                "php://filter/convert.base64-encode/resource=index.php"
            ],
            "usage": "These payloads should be inserted into file inclusion parameters, such as ?page=, ?include=, or ?file=.",
            "risk_level": "high"
        }
    
    def _generate_rce_payload(self, finding):
        """Generate RCE payloads"""
        return {
            "type": "rce",
            "name": "Remote Code Execution Payload",
            "target": finding.get("host", self.target),
            "port": finding.get("port", "80"),
            "vulnerability": "Remote Code Execution",
            "description": "Payloads to exploit command injection vulnerabilities",
            "payloads": [
                "; id",
                "| id",
                "$(id)",
                "`id`",
                "'; id",
                "\"; id",
                "& id",
                "&& id"
            ],
            "usage": "These payloads should be inserted into command execution points.",
            "risk_level": "critical"
        }
    
    def identify_attack_vectors(self, exploit_filter=None):
        """Identify potential attack vectors based on findings"""
        logger.info("Identifying attack vectors")
        
        attack_vectors = []
        
        # Process validated vulnerabilities
        if "vulnerability_validation" in self.results and "vulnerabilities" in self.results["vulnerability_validation"]:
            vulns = self.results["vulnerability_validation"]["vulnerabilities"]
            
            for vuln in vulns:
                risk_level = vuln.get("risk_level")
                
                if exploit_filter and "risk_level" in exploit_filter:
                    if risk_level not in exploit_filter["risk_level"]:
                        continue
                
                attack_vectors.append({
                    "type": "vulnerability",
                    "name": vuln.get("vulnerability"),
                    "target": f"{vuln.get('host')}:{vuln.get('port')}",
                    "service": vuln.get("service"),
                    "risk_level": risk_level,
                    "description": vuln.get("details", "No details available"),
                    "payload_available": any(p["type"] == vuln.get("vulnerability", "").lower() for p in self.results.get("payload_generation", {}).get("payloads", []))
                })
        
        # Process web app findings
        if "webapp_analysis" in self.results and "findings" in self.results["webapp_analysis"]:
            findings = self.results["webapp_analysis"]["findings"]
            
            for finding in findings:
                risk_level = finding.get("risk_level")
                
                if exploit_filter and "risk_level" in exploit_filter:
                    if risk_level not in exploit_filter["risk_level"]:
                        continue
                
                attack_vectors.append({
                    "type": "web",
                    "name": finding.get("finding"),
                    "target": finding.get("url", f"{finding.get('host')}:{finding.get('port')}"),
                    "category": finding.get("category"),
                    "risk_level": risk_level,
                    "description": finding.get("description", "No description available"),
                    "payload_available": any(p["type"] == finding.get("category", "").lower() for p in self.results.get("payload_generation", {}).get("payloads", []))
                })
        
        # Process credential findings
        if "credential_testing" in self.results and "findings" in self.results["credential_testing"]:
            findings = self.results["credential_testing"]["findings"]
            
            for finding in findings:
                if "credentials" in finding:
                    attack_vectors.append({
                        "type": "credentials",
                        "name": finding.get("finding"),
                        "target": f"{finding.get('host')}:{finding.get('port')}",
                        "service": finding.get("service"),
                        "risk_level": finding.get("risk_level"),
                        "description": finding.get("description", "No description available"),
                        "credentials": finding.get("credentials")
                    })
        
        # Sort attack vectors by risk level
        risk_order = {"critical": 1, "high": 2, "medium": 3, "low": 4, "info": 5}
        attack_vectors.sort(key=lambda x: risk_order.get(x["risk_level"], 99))
        
        self.results["attack_vectors"] = attack_vectors
        logger.info(f"Identified {len(attack_vectors)} potential attack vectors")
        return attack_vectors
    
    def generate_report(self, format="all"):
        """Generate report in specified format"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_safe = self.target.replace(".", "_").replace(":", "_")
        base_filename = f"{self.output_dir}/{target_safe}_preexploit_{timestamp}"
        
        report_files = {}
        
        if format in ["json", "all"]:
            json_file = f"{base_filename}.json"
            self._generate_json_report(json_file)
            report_files["json"] = json_file
            
        if format in ["excel", "all"]:
            if EXCEL_AVAILABLE:
                excel_file = f"{base_filename}.xlsx"
                self._generate_excel_report(excel_file)
                report_files["excel"] = excel_file
            else:
                logger.warning("Excel report skipped - openpyxl library not available")
                report_files["excel"] = None
            
        if format in ["md", "all"]:
            md_file = f"{base_filename}.md"
            self._generate_md_report(md_file)
            report_files["markdown"] = md_file
            
        logger.info(f"Reports generated in {self.output_dir} directory")
        return report_files
    
    def _generate_json_report(self, filename):
        """Generate JSON report"""
        with open(filename, "w") as f:
            json.dump(self.results, f, indent=4)
    
    def _generate_excel_report(self, filename):
        """Generate Excel report"""
        if not EXCEL_AVAILABLE:
            logger.error("Cannot generate Excel report - openpyxl library not available")
            return
        
        wb = Workbook()
        
        # Summary sheet
        summary = wb.active
        summary.title = "Summary"
        summary['A1'] = "Pre-Exploitation Report"
        summary['A1'].font = Font(bold=True, size=14)
        summary['A3'] = "Target:"
        summary['B3'] = self.target
        summary['A4'] = "Timestamp:"
        summary['B4'] = self.results["timestamp"]
        
        # Add vulnerability summary
        if "vulnerability_validation" in self.results:
            vuln_count = self.results["vulnerability_validation"].get("total_count", 0)
            summary['A6'] = "Vulnerabilities Summary"
            summary['A6'].font = Font(bold=True)
            summary['A7'] = "Total Vulnerabilities:"
            summary['B7'] = vuln_count
            
            # Count by risk level
            risk_levels = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
            for vuln in self.results["vulnerability_validation"].get("vulnerabilities", []):
                risk = vuln.get("risk_level", "info").lower()
                if risk in risk_levels:
                    risk_levels[risk] += 1
            
            row = 8
            for risk, count in risk_levels.items():
                summary[f'A{row}'] = f"{risk.capitalize()} Risk:"
                summary[f'B{row}'] = count
                if count > 0 and self.risk_colors.get(risk):
                    summary[f'B{row}'].fill = self.risk_colors[risk]
                row += 1
        
        # Add web findings summary
        if "webapp_analysis" in self.results:
            web_count = self.results["webapp_analysis"].get("total_findings", 0)
            summary['A11'] = "Web Findings Summary"
            summary['A11'].font = Font(bold=True)
            summary['A12'] = "Total Web Findings:"
            summary['B12'] = web_count
            
            # Count by category
            if "grouped_findings" in self.results["webapp_analysis"]:
                row = 13
                for category, findings in self.results["webapp_analysis"]["grouped_findings"].items():
                    summary[f'A{row}'] = f"{category}:"
                    summary[f'B{row}'] = len(findings)
                    row += 1
        
        # Add attack vectors
        if "attack_vectors" in self.results:
            attack_count = len(self.results["attack_vectors"])
            summary['A15'] = "Attack Vectors Summary"
            summary['A15'].font = Font(bold=True)
            summary['A16'] = "Total Attack Vectors:"
            summary['B16'] = attack_count
            
            # List top attack vectors
            if attack_count > 0:
                summary['A18'] = "Top Attack Vectors"
                summary['A18'].font = Font(bold=True)
                
                summary['A19'] = "Vector"
                summary['B19'] = "Target"
                summary['C19'] = "Type"
                summary['D19'] = "Risk Level"
                
                row = 20
                for i, vector in enumerate(self.results["attack_vectors"][:10]):  # Show top 10
                    summary[f'A{row}'] = vector.get("name", "Unknown")
                    summary[f'B{row}'] = vector.get("target", "Unknown")
                    summary[f'C{row}'] = vector.get("type", "Unknown")
                    summary[f'D{row}'] = vector.get("risk_level", "Unknown")
                    
                    risk = vector.get("risk_level", "").lower()
                    if risk in self.risk_colors:
                        summary[f'D{row}'].fill = self.risk_colors[risk]
                    
                    row += 1
        
        # Vulnerabilities sheet
        if "vulnerability_validation" in self.results and "vulnerabilities" in self.results["vulnerability_validation"]:
            vulns_sheet = wb.create_sheet("Vulnerabilities")
            vulns_sheet['A1'] = "Validated Vulnerabilities"
            vulns_sheet['A1'].font = Font(bold=True, size=14)
            
            vulns_sheet['A3'] = "Name"
            vulns_sheet['B3'] = "Host"
            vulns_sheet['C3'] = "Port"
            vulns_sheet['D3'] = "Service"
            vulns_sheet['E3'] = "Risk Level"
            vulns_sheet['F3'] = "CVE"
            vulns_sheet['G3'] = "Details"
            
            row = 4
            for vuln in self.results["vulnerability_validation"]["vulnerabilities"]:
                vulns_sheet[f'A{row}'] = vuln.get("vulnerability", "Unknown")
                vulns_sheet[f'B{row}'] = vuln.get("host", "Unknown")
                vulns_sheet[f'C{row}'] = vuln.get("port", "Unknown")
                vulns_sheet[f'D{row}'] = vuln.get("service", "Unknown")
                vulns_sheet[f'E{row}'] = vuln.get("risk_level", "Unknown")
                vulns_sheet[f'F{row}'] = vuln.get("cve", "N/A")
                vulns_sheet[f'G{row}'] = vuln.get("details", "No details available")
                
                risk = vuln.get("risk_level", "").lower()
                if risk in self.risk_colors:
                    vulns_sheet[f'E{row}'].fill = self.risk_colors[risk]
                
                row += 1
        
        # Web Findings sheet
        if "webapp_analysis" in self.results and "findings" in self.results["webapp_analysis"]:
            web_sheet = wb.create_sheet("Web Findings")
            web_sheet['A1'] = "Web Application Findings"
            web_sheet['A1'].font = Font(bold=True, size=14)
            
            web_sheet['A3'] = "Finding"
            web_sheet['B3'] = "URL"
            web_sheet['C3'] = "Category"
            web_sheet['D3'] = "Risk Level"
            web_sheet['E3'] = "Description"
            web_sheet['F3'] = "Recommendation"
            
            row = 4
            for finding in self.results["webapp_analysis"]["findings"]:
                web_sheet[f'A{row}'] = finding.get("finding", "Unknown")
                host_port = f"{finding.get('host', '')}:{finding.get('port', '')}"
                web_sheet[f'B{row}'] = finding.get("url", host_port)
                web_sheet[f'C{row}'] = finding.get("category", "Unknown")
                web_sheet[f'D{row}'] = finding.get("risk_level", "Unknown")
                web_sheet[f'E{row}'] = finding.get("description", "No description available")
                web_sheet[f'F{row}'] = finding.get("recommendation", "No recommendation available")
                
                risk = finding.get("risk_level", "").lower()
                if risk in self.risk_colors:
                    web_sheet[f'D{row}'].fill = self.risk_colors[risk]
                
                row += 1
        
        # Adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(filename)
    
    def _generate_md_report(self, filename):
        """Generate Markdown report"""
        with open(filename, "w") as f:
            f.write(f"# Pre-Exploitation Report\n\n")
            f.write(f"## Target: {self.target}\n")
            f.write(f"## Timestamp: {self.results['timestamp']}\n\n")
            
            # Summary
            if "vulnerability_validation" in self.results:
                vuln_count = self.results["vulnerability_validation"].get("total_count", 0)
                f.write(f"## Summary\n\n")
                f.write(f"- **Total Vulnerabilities:** {vuln_count}\n")
                
                # Count by risk level
                risk_levels = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
                for vuln in self.results["vulnerability_validation"].get("vulnerabilities", []):
                    risk = vuln.get("risk_level", "info").lower()
                    if risk in risk_levels:
                        risk_levels[risk] += 1
                
                for risk, count in risk_levels.items():
                    if count > 0:
                        f.write(f"  - {risk.capitalize()}: {count}\n")
                
                f.write("\n")
            
            # Web findings
            if "webapp_analysis" in self.results:
                web_count = self.results["webapp_analysis"].get("total_findings", 0)
                f.write(f"- **Web Application Findings:** {web_count}\n")
                
                # Categories
                if "grouped_findings" in self.results["webapp_analysis"]:
                    for category, findings in self.results["webapp_analysis"]["grouped_findings"].items():
                        f.write(f"  - {category}: {len(findings)}\n")
                
                f.write("\n")
            
            # Attack vectors
            if "attack_vectors" in self.results:
                attack_count = len(self.results["attack_vectors"])
                f.write(f"- **Attack Vectors:** {attack_count}\n\n")
            
            # Vulnerabilities
            if "vulnerability_validation" in self.results and "vulnerabilities" in self.results["vulnerability_validation"]:
                f.write(f"## Validated Vulnerabilities\n\n")
                
                # Group by risk level
                for risk in ["critical", "high", "medium", "low", "info"]:
                    risk_vulns = [v for v in self.results["vulnerability_validation"]["vulnerabilities"] if v.get("risk_level") == risk]
                    
                    if risk_vulns:
                        f.write(f"### {risk.capitalize()} Risk Vulnerabilities\n\n")
                        
                        for vuln in risk_vulns:
                            f.write(f"#### {vuln.get('vulnerability', 'Unknown')}\n\n")
                            f.write(f"- **Host:** {vuln.get('host', 'Unknown')}\n")
                            f.write(f"- **Port:** {vuln.get('port', 'Unknown')}\n")
                            f.write(f"- **Service:** {vuln.get('service', 'Unknown')}\n")
                            
                            if "cve" in vuln:
                                f.write(f"- **CVE:** {vuln['cve']}\n")
                            
                            if "details" in vuln:
                                f.write(f"- **Details:** {vuln['details']}\n")
                            elif "description" in vuln:
                                f.write(f"- **Description:** {vuln['description']}\n")
                            
                            f.write("\n")
            
            # Web findings
            if "webapp_analysis" in self.results and "findings" in self.results["webapp_analysis"]:
                f.write(f"## Web Application Findings\n\n")
                
                # Group by category
                if "grouped_findings" in self.results["webapp_analysis"]:
                    for category, findings in self.results["webapp_analysis"]["grouped_findings"].items():
                        f.write(f"### {category}\n\n")
                        
                        for finding in findings:
                            f.write(f"#### {finding.get('finding', 'Unknown')}\n\n")
                            host_port = f"{finding.get('host', '')}:{finding.get('port', '')}"
                            f.write(f"- **URL:** {finding.get('url', host_port)}\n")
                            f.write(f"- **Risk Level:** {finding.get('risk_level', 'Unknown')}\n")
                            f.write(f"- **Description:** {finding.get('description', 'No description available')}\n")
                            f.write(f"- **Recommendation:** {finding.get('recommendation', 'No recommendation available')}\n\n")
                else:
                    # Fallback to flat list
                    for finding in self.results["webapp_analysis"]["findings"]:
                        f.write(f"### {finding.get('finding', 'Unknown')}\n\n")
                        host_port = f"{finding.get('host', '')}:{finding.get('port', '')}"
                        f.write(f"- **URL:** {finding.get('url', host_port)}\n")
                        f.write(f"- **Category:** {finding.get('category', 'Unknown')}\n")
                        f.write(f"- **Risk Level:** {finding.get('risk_level', 'Unknown')}\n")
                        f.write(f"- **Description:** {finding.get('description', 'No description available')}\n")
                        f.write(f"- **Recommendation:** {finding.get('recommendation', 'No recommendation available')}\n\n")
            
            # Attack vectors
            if "attack_vectors" in self.results:
                f.write(f"## Attack Vectors\n\n")
                
                f.write(f"| Vector | Target | Type | Risk Level |\n")
                f.write(f"|--------|--------|------|------------|\n")
                
                for vector in self.results["attack_vectors"]:
                    f.write(f"| {vector.get('name', 'Unknown')} | {vector.get('target', 'Unknown')} | {vector.get('type', 'Unknown')} | {vector.get('risk_level', 'Unknown')} |\n")
                
                f.write("\n")
            
            # Final section
            f.write(f"## Recommendations\n\n")
            
            # Extract recommendations from findings
            recommendations = set()
            
            if "webapp_analysis" in self.results and "findings" in self.results["webapp_analysis"]:
                for finding in self.results["webapp_analysis"]["findings"]:
                    if "recommendation" in finding and finding.get("risk_level") in ["critical", "high", "medium"]:
                        recommendations.add(finding["recommendation"])
            
            # Add generic recommendations
            if not recommendations:
                recommendations = {
                    "Implement proper input validation for all user inputs",
                    "Enable security headers on web servers",
                    "Regularly update and patch software",
                    "Implement strong password policies",
                    "Remove or secure default accounts"
                }
            
            # Write recommendations
            for rec in recommendations:
                f.write(f"- {rec}\n")
            
            # End of report
            f.write(f"\n*Report generated on {self.results['timestamp']}*\n")

class ExploitationPhase:
    """Exploitation Phase Controller"""
    
    def __init__(self, framework):
        """Initialize the Exploitation phase controller"""
        self.framework = framework
        self.target = framework.target
        self.output_dir = framework.output_dir
        self.threads = framework.threads
        self.preexploit_results = framework.results.get("preexploit", {})
        self.results = {
            "target": self.target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "exploitation_summary": {
                "attempts": 0,
                "successful": 0,
                "failed": 0
            },
            "exploits": [],
            "shells": [],
            "privilege_escalation": []
        }
    
    def run(self, auto_exploit=False, exploit_filter=None):
        """Run exploitation phase"""
        logger.info(f"Starting exploitation phase for {self.target}")
        
        # This is a placeholder for the exploitation phase
        # In a real implementation, would perform actual exploitation
        
        return self.results

class PostExploitationPhase:
    """Post-Exploitation Phase Controller"""
    
    def __init__(self, framework):
        """Initialize the Post-Exploitation phase controller"""
        self.framework = framework
        self.target = framework.target
        self.output_dir = framework.output_dir
        self.threads = framework.threads
        self.exploit_results = framework.results.get("exploit", {})
        self.results = {
            "target": self.target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "persistence": [],
            "data_exfiltration": [],
            "evidence_removal": []
        }
    
    def run(self, install_persistence=False, exfiltrate_data=False, cleanup_traces=True):
        """Run post-exploitation phase"""
        logger.info(f"Starting post-exploitation phase for {self.target}")
        
        # This is a placeholder for the post-exploitation phase
        # In a real implementation, would perform actual post-exploitation activities
        
        return self.results

def parse_arguments():
    """Parse command-line arguments"""
    parser = argparse.ArgumentParser(description='APTES - Advanced Penetration Testing and Exploitation Suite')
    
    # Target argument
    parser.add_argument('target', nargs='?', help='Target host or IP address')

    
    # Phase selection
    parser.add_argument('-p', '--phase', choices=['recon', 'preexploit', 'exploit', 'postexploit', 'all'],
                        default='preexploit', help='Phase to run (default: preexploit)')
    
    # General options
    parser.add_argument('-o', '--output-dir', default='reports',
                        help='Directory for output reports (default: reports)')
    
    parser.add_argument('-r', '--results-file',
                        help='Load results from previous phase')
    
    parser.add_argument('--threads', type=int, default=3,
                        help='Number of threads for concurrent operations (default: 3)')
    
    parser.add_argument('--format', choices=['json', 'excel', 'md', 'all'], default='all',
                        help='Report format (default: all)')
    
    # Reconnaissance phase options
    parser.add_argument('--passive-only', action='store_true',
                        help='Perform only passive reconnaissance')
    
    parser.add_argument('--skip-web', action='store_true',
                        help='Skip web scanning during recon')
    
    # Pre-exploitation phase options
    parser.add_argument('--filter', choices=['all', 'critical', 'high', 'medium', 'low'],
                        default='all', help='Risk level filter (default: all)')
    
    parser.add_argument('--no-creds', action='store_true',
                        help='Skip default credential testing')
    
    parser.add_argument('--no-payloads', action='store_true',
                        help='Skip payload generation')
    
    # Exploitation phase options
    parser.add_argument('--auto-exploit', action='store_true',
                        help='Automatically exploit without confirmation')
    
    # Post-exploitation phase options
    parser.add_argument('--install-persistence', action='store_true',
                        help='Install persistence mechanisms')
    
    parser.add_argument('--exfiltrate-data', action='store_true',
                        help='Exfiltrate sensitive data')
    
    parser.add_argument('--no-cleanup', action='store_true',
                        help='Skip cleaning up traces of activity')
    
    # Verbosity options
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Enable verbose output')
    
    parser.add_argument('-q', '--quiet', action='store_true',
                        help='Suppress all output except errors')
    
    parser.add_argument('--no-verify-ssl', action='store_true',
                    help='Disable SSL certificate verification (security risk)')
    
    parser.add_argument('--suppress-ssl-warnings', action='store_true', default=True,
                        help='Suppress SSL certificate warnings')
    
    args = parser.parse_args()
    
    # Handle custom help flag
    if len(sys.argv) == 1:
        print_usage()
        sys.exit(1)
        
    return args

def print_usage():
    """Print usage information for APTES"""
    print("\nAPTES - Advanced Penetration Testing and Exploitation Suite")
    print("\nUsage: python aptes.py [TARGET] [OPTIONS]")
    print("\nError: No target specified. Please provide a target to scan.")
    
    print("\nBasic Usage:")
    print("  python aptes.py <target> [options]")
    print("\nExamples:")
    print("  python aptes.py example.com                # Run pre-exploitation phase against example.com")
    print("  python aptes.py 192.168.1.10 -p recon      # Run only reconnaissance phase")
    print("  python aptes.py target.org --filter high   # Focus on high-risk vulnerabilities")
    
    print("\nOptions:")
    print("  Target Selection:")
    print("    TARGET                  Target host or IP address to assess")
    print("\n  Phase Selection:")
    print("    -p, --phase {recon,preexploit,exploit,postexploit,all}")
    print("                            Phase to run (default: preexploit)")
    
    print("\n  General Options:")
    print("    -o, --output-dir DIR    Directory for output reports (default: reports)")
    print("    -r, --results-file FILE Load results from previous phase")
    print("    --threads N             Number of threads for concurrent operations (default: 3)")
    print("    --format {json,excel,md,all}")
    print("                            Report format (default: all)")
    
    # Add all other option categories here...
    
    print("\n  Verbosity Options:")
    print("    -v, --verbose           Enable verbose output")
    print("    -q, --quiet             Suppress all output except errors")
    print("    -h, --help              Show this help message and exit")
    
    print("\nFor full documentation, visit: https://github.com/byteshell/aptes")

def print_summary(framework, phase):
    """Print a summary of results from a phase"""
    # Safely get results for the phase
    results = {}
    try:
        results = framework.results.get(phase, {})
    except (AttributeError, KeyError):
        print(f"\nNo results available for {phase} phase")
        return
    
    if not results:
        print(f"\nNo results available for {phase} phase")
        return
    
    print(f"\n{'=' * 50}")
    print(f"{phase.upper()} PHASE SUMMARY")
    print(f"{'=' * 50}\n")
    
    if phase == "recon":
        # Print reconnaissance summary
        if "passive" in results:
            print("Passive Reconnaissance:")
            if "dns" in results["passive"]:
                print(f"  - DNS Records: {sum(len(records) for records in results['passive']['dns'].values())} entries found")
            if "subdomains" in results["passive"]:
                print(f"  - Subdomains: {len(results['passive']['subdomains'])} discovered")
            if "ssl_info" in results["passive"]:
                print(f"  - SSL Information: {'Collected' if results['passive']['ssl_info'] else 'None'}")
        
        if "active" in results:
            print("\nActive Reconnaissance:")
            if "ports" in results["active"]:
                open_ports = 0
                for host_data in results["active"]["ports"].values():
                    for proto_data in host_data.values():
                        open_ports += len(proto_data)
                print(f"  - Open Ports: {open_ports} discovered")
            if "services" in results["active"]:
                services = set()
                for host_data in results["active"]["services"].values():
                    for proto_data in host_data.values():
                        for port_data in proto_data.values():
                            services.add(port_data.get("service", "unknown"))
                print(f"  - Services: {', '.join(services) if services else 'None identified'}")
            if "vulnerabilities" in results["active"]:
                vulns = results["active"]["vulnerabilities"].get("vulnerabilities", [])
                print(f"  - Potential Vulnerabilities: {len(vulns)} found")
            if "web" in results["active"]:
                web_results = results["active"]["web"]
                print(f"  - Web Technologies: {', '.join(web_results.get('technologies', [])) if web_results.get('technologies') else 'None identified'}")
    
    elif phase == "preexploit":
        # Print pre-exploitation summary
        if "vulnerability_validation" in results:
            vuln_count = results["vulnerability_validation"].get("total_count", 0)
            print(f"Validated Vulnerabilities: {vuln_count}")
            
            # Count by risk level
            risk_levels = {"critical": 0, "high": 0, "medium": 0, "low": 0, "info": 0}
            for vuln in results["vulnerability_validation"].get("vulnerabilities", []):
                risk = vuln.get("risk_level", "info").lower()
                if risk in risk_levels:
                    risk_levels[risk] += 1
            
            for risk, count in risk_levels.items():
                if count > 0:
                    print(f"  - {risk.capitalize()}: {count}")
        
        if "webapp_analysis" in results:
            web_count = results["webapp_analysis"].get("total_findings", 0)
            print(f"\nWeb Application Findings: {web_count}")
            
            # Categories
            if "grouped_findings" in results["webapp_analysis"]:
                for category, findings in results["webapp_analysis"]["grouped_findings"].items():
                    print(f"  - {category}: {len(findings)}")
        
        if "attack_vectors" in results:
            attack_count = len(results["attack_vectors"])
            print(f"\nIdentified Attack Vectors: {attack_count}")
            
            # Show top 3 attack vectors
            top_vectors = sorted([v for v in results["attack_vectors"] 
                               if v.get("risk_level") in ["critical", "high"]], 
                              key=lambda x: {"critical": 0, "high": 1}.get(x.get("risk_level"), 2))[:3]
            
            if top_vectors:
                print(f"\nTop attack vectors:")
                for i, vector in enumerate(top_vectors, 1):
                    print(f"  {i}. [{vector.get('risk_level', '').upper()}] {vector.get('name', 'Unknown')}")
    
    elif phase == "exploit":
        # Print exploitation summary
        print("Exploitation Summary:")
        print(f"  - Attempts: {results['exploitation_summary'].get('attempts', 0)}")
        print(f"  - Successful: {results['exploitation_summary'].get('successful', 0)}")
        print(f"  - Failed: {results['exploitation_summary'].get('failed', 0)}")
        
        if "shells" in results and results["shells"]:
            print(f"\nObtained Shells: {len(results['shells'])}")
            for shell in results["shells"]:
                print(f"  - {shell.get('type', 'Unknown')} shell on {shell.get('target', 'Unknown')} ({shell.get('privileges', 'unknown')} privileges)")
    
    elif phase == "postexploit":
        # Print post-exploitation summary
        if "persistence" in results:
            print(f"Persistence Mechanisms: {len(results['persistence'])}")
            for mechanism in results["persistence"]:
                print(f"  - {mechanism.get('technique', 'Unknown')} on {mechanism.get('host', 'Unknown')}")
        
        if "data_exfiltration" in results:
            print(f"\nData Exfiltration:")
            total_data = sum(op.get("total_size", 0) for op in results["data_exfiltration"])
            print(f"  - Total data exfiltrated: {total_data / (1024*1024):.2f} MB")
            
            data_types = set()
            for op in results["data_exfiltration"]:
                for data_type in op.get("data_types", []):
                    data_types.add(data_type)
            
            if data_types:
                print(f"  - Types of data: {', '.join(data_types)}")
    
    print(f"\n{'=' * 50}")

def interactive_target_selection():
    """Interactively prompt for a target if not provided"""
    print("\nTarget not specified. Please enter a target.")
    while True:
        target = input("Enter target host or IP address: ").strip()
        if target:
            return target
        print("Target is required. Please try again.")

def main():
    """Main function"""
    # Parse command-line arguments
    args = parse_arguments()
    
    # Interactively prompt for target if not provided
    if not args.target:
        print_usage()
        return 1
    
    # Set verbosity level
    verbosity = 0
    if args.quiet:
        verbosity = 0
    elif args.verbose:
        verbosity = 2
    else:
        verbosity = 1

    if args.suppress_ssl_warnings:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    verify_ssl = not args.no_verify_ssl
    
    # Initialize the APTES framework
    framework = APTESFramework(
        target=args.target,
        output_dir=args.output_dir,
        threads=args.threads,
        verbosity=verbosity,
        verify_ssl=verify_ssl
    )
    
    # Print banner
    if verbosity > 0:
        framework.print_banner()
    
    # Load previous results if provided
    if args.results_file:
        if not framework.load_results(args.results_file):
            logger.error(f"Failed to load results from {args.results_file}")
            return 1
    
    # Build risk level filter
    exploit_filter = None
    if args.filter != 'all':
        exploit_filter = {'risk_level': [args.filter]}
        if args.filter == 'critical':
            exploit_filter['risk_level'] = ['critical']
        elif args.filter == 'high':
            exploit_filter['risk_level'] = ['critical', 'high']
        elif args.filter == 'medium':
            exploit_filter['risk_level'] = ['critical', 'high', 'medium']
        elif args.filter == 'low':
            exploit_filter['risk_level'] = ['critical', 'high', 'medium', 'low']
    
    # Determine which phases to run
    phases = []
    if args.phase == 'all':
        phases = ['recon', 'preexploit', 'exploit', 'postexploit']
    else:
        phases = [args.phase]
    
    # Run the specified phases
    for phase in phases:
        logger.info(f"Starting {phase} phase")
        
        try:
            if phase == 'recon':
                # Run reconnaissance phase
                success = framework.run_phase('recon', 
                                             passive_only=args.passive_only,
                                             skip_web=args.skip_web)
            
            elif phase == 'preexploit':
                # Run pre-exploitation phase
                success = framework.run_phase('preexploit',
                                             exploit_filter=exploit_filter,
                                             skip_web=args.skip_web,
                                             skip_creds=args.no_creds,
                                             skip_payloads=args.no_payloads)
            
            elif phase == 'exploit':
                # Run exploitation phase
                success = framework.run_phase('exploit',
                                             auto_exploit=args.auto_exploit,
                                             exploit_filter=exploit_filter)
            
            elif phase == 'postexploit':
                # Run post-exploitation phase
                success = framework.run_phase('postexploit',
                                             install_persistence=args.install_persistence,
                                             exfiltrate_data=args.exfiltrate_data,
                                             cleanup_traces=not args.no_cleanup)
            
            if not success:
                logger.error(f"Failed to run {phase} phase")
                if phase != phases[-1]:  # Only exit if not the last phase
                    return 1
            
            # Print summary
            if verbosity > 0:
                print_summary(framework, phase)
                
            # Generate reports
            try:
                if phase == 'preexploit':
                    # Generate reports for pre-exploitation phase
                    phase_controller = getattr(framework, phase)
                    if hasattr(phase_controller, 'generate_report'):
                        report_files = phase_controller.generate_report(format=args.format)
                        logger.info(f"Generated reports: {', '.join([f for f in report_files.values() if f])}")
                    else:
                        # Fallback to direct JSON report
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        target_safe = args.target.replace(".", "_").replace(":", "_")
                        json_file = f"{args.output_dir}/{target_safe}_preexploit_{timestamp}.json"
                        
                        with open(json_file, "w") as f:
                            json.dump(framework.results.get(phase, {}), f, indent=4)
                        logger.info(f"Generated JSON report: {json_file}")
            except Exception as e:
                logger.error(f"Error generating reports: {str(e)}")
            
        except KeyboardInterrupt:
            logger.error(f"{phase.capitalize()} phase interrupted by user")
            return 1
        except Exception as e:
            logger.error(f"Error in {phase} phase: {str(e)}")
            if verbosity >= 2:
                import traceback
                traceback.print_exc()
            return 1
    
    # Save final results
    try:
        results_file = framework.save_results()
        logger.info(f"Results saved to {results_file}")
    except Exception as e:
        logger.error(f"Error saving results: {str(e)}")
        # Create a fallback save method
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_safe = args.target.replace(".", "_").replace(":", "_")
        fallback_file = f"{args.output_dir}/{target_safe}_results_{timestamp}.json"
        
        try:
            with open(fallback_file, "w") as f:
                json.dump(framework.results, f, indent=4)
            logger.info(f"Results saved to fallback file: {fallback_file}")
        except Exception:
            logger.error("Failed to save results to fallback file")
    
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nOperation interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {str(e)}")
        if '--verbose' in sys.argv or '-v' in sys.argv:
            import traceback
            traceback.print_exc()
        sys.exit(1)
