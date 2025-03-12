import os
import re
import json
import requests
import subprocess
import socket
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
import concurrent.futures

class Reconnaissance:
    def __init__(self, target, output_dir="reports"):
        """
        Initialize the reconnaissance module with a target and output directory
        """
        self.target = target
        self.output_dir = output_dir
        self.results = {
            "target": target,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "passive": {},
            "active": {}
        }
        
        self.outdated_services = {
            "services": [],
            "recommendations": []
        }
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
    
    def run_full_scan(self):
        """Run both passive and active reconnaissance"""
        self.passive_recon()
        self.active_recon()
        self.check_outdated_services()
        return self.results
    
    def passive_recon(self):
        """Perform passive reconnaissance"""
        print(f"[*] Starting passive reconnaissance on {self.target}")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            dns_future = executor.submit(self.dns_lookup)
            subdomain_future = executor.submit(self.subdomain_enum)
            ssl_future = executor.submit(self.ssl_info)
            
            self.results["passive"]["dns"] = dns_future.result()
            self.results["passive"]["subdomains"] = subdomain_future.result()
            self.results["passive"]["ssl_info"] = ssl_future.result()
            
        print(f"[+] Passive reconnaissance completed")
        return self.results["passive"]
    
    def active_recon(self):
        """Perform active reconnaissance"""
        print(f"[*] Starting active reconnaissance on {self.target}")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            port_scan_future = executor.submit(self.port_scan)
            service_enum_future = executor.submit(self.service_enum)
            vuln_scan_future = executor.submit(self.vuln_scan)
            
            self.results["active"]["ports"] = port_scan_future.result()
            self.results["active"]["services"] = service_enum_future.result()
            self.results["active"]["vulnerabilities"] = vuln_scan_future.result()
        
        self.results["active"]["web"] = self.web_scan()
        
        print(f"[+] Active reconnaissance completed")
        return self.results["active"]
    
    def dns_lookup(self):
        """Perform DNS lookups using subprocess"""
        print(f"[*] Performing DNS lookups for {self.target}")
        record_types = ["A", "AAAA", "MX", "NS", "TXT", "SOA"]
        results = {}
        
        for record_type in record_types:
            try:
                cmd = f"dig +short {self.target} {record_type}"
                output = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
                
                if output:
                    results[record_type] = output.split('\n')
                else:
                    results[record_type] = []
            except Exception:
                results[record_type] = []
                
        return results
    
    def subdomain_enum(self):
        """Enumerate subdomains using public sources"""
        print(f"[*] Enumerating subdomains for {self.target}")
        subdomains = set()
        
        try:
            response = requests.get(f"https://crt.sh/?q=%.{self.target}&output=json")
            if response.status_code == 200:
                data = response.json()
                for entry in data:
                    domain_name = entry['name_value']
                    subdomains.add(domain_name)
        except Exception as e:
            print(f"[!] Error in crt.sh lookup: {str(e)}")
        
        try:
            cmd = f"subfinder -d {self.target} -silent"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
            if output:
                for subdomain in output.split('\n'):
                    subdomains.add(subdomain.strip())
        except Exception:
            pass
            
        return list(subdomains)
    
    def ssl_info(self):
        """Gather SSL certificate information"""
        print(f"[*] Gathering SSL certificate information for {self.target}")
        try:
            cmd = f"echo | openssl s_client -servername {self.target} -connect {self.target}:443 2>/dev/null | openssl x509 -noout -text"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            result = {}
            issuer_match = re.search(r"Issuer:.*?=(.*?)[,\n]", output)
            if issuer_match:
                result["issuer"] = issuer_match.group(1).strip()
                
            expiry_match = re.search(r"Not After\s*:\s*(.*?)\n", output)
            if expiry_match:
                result["expiry"] = expiry_match.group(1).strip()
                
            subject_alt_match = re.search(r"Subject Alternative Name:.*?DNS:(.*?)(?:,|$)", output)
            if subject_alt_match:
                result["alt_names"] = subject_alt_match.group(1).strip()
                
            return result
        except Exception as e:
            return {"error": str(e)}
    
    def port_scan(self):
        """Perform port scanning using nmap via subprocess"""
        print(f"[*] Scanning ports on {self.target}")
        try:
            cmd = f"nmap -F -T4 {self.target} -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            return results
        except Exception as e:
            return {"error": str(e)}
    
    def resolve_domain_to_ip(self, domain):
        """Resolve a domain name to its IP address"""
        try:
            return socket.gethostbyname(domain)
        except socket.gaierror:
            return domain  
    
    def service_enum(self):
        """Enumerate services on open ports using nmap via subprocess"""
        print(f"[*] Enumerating services on {self.target}")
        try:
            target_ip = self.resolve_domain_to_ip(self.target)
            print(f"[*] Resolved {self.target} to {target_ip}")
            
            cmd = f"nmap -sV -T4 --script=banner {target_ip} -p 21,22,25,80,443,8080,8443 -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            return results
        except Exception as e:
            return {"error": str(e)}
    
    def vuln_scan(self):
        """Perform basic vulnerability scanning using nmap scripts"""
        print(f"[*] Checking for common vulnerabilities on {self.target}")
        try:
            target_ip = self.resolve_domain_to_ip(self.target)
            
            cmd = f"nmap -sV --script=vuln {target_ip} -p 21,22,25,80,443,8080,8443 -oX -"
            output = subprocess.check_output(cmd, shell=True).decode('utf-8')
            
            results = self._parse_nmap_xml(output)
            
            vulns = []
            for host_data in results.values():
                for proto_data in host_data.values():
                    for port_data in proto_data.values():
                        scripts = port_data.get("scripts", {})
                        for script_name, script_output in scripts.items():
                            if "VULNERABLE" in script_output:
                                vulns.append({
                                    "port": port_data.get("port", "unknown"),
                                    "service": port_data.get("service", "unknown"),
                                    "vulnerability": script_name,
                                    "details": script_output
                                })
            
            return {"vulnerabilities": vulns, "scan_data": results}
        except Exception as e:
            return {"error": str(e)}
    
    def check_outdated_services(self):
        """Check for outdated services and provide update recommendations"""
        print(f"[*] Checking for outdated services")
        
        latest_versions = {
            "apache": {"version": "2.4.57", "recommended": "Apache 2.4.57 or later"},
            "nginx": {"version": "1.24.0", "recommended": "NGINX 1.24.0 or later"},
            "openssh": {"version": "9.5", "recommended": "OpenSSH 9.5 or later"},
            "php": {"version": "8.3.0", "recommended": "PHP 8.3 or later"},
            "mysql": {"version": "8.0.35", "recommended": "MySQL 8.0.35 or later"},
            "mariadb": {"version": "11.2.3", "recommended": "MariaDB 11.2.3 or later"},
            "postgresql": {"version": "16.1", "recommended": "PostgreSQL 16.1 or later"},
            "vsftpd": {"version": "3.0.5", "recommended": "VSFTPD 3.0.5 or later"},
            "proftpd": {"version": "1.3.8", "recommended": "ProFTPD 1.3.8 or later"}
        }
        
        def extract_version(product_str, version_str):
            """Extract clean version numbers from product and version strings"""
            combined = f"{product_str} {version_str}".lower()
            
            for service_name in latest_versions.keys():
                if service_name in combined:
                    version_match = re.search(r'(\d+\.\d+\.?\d*)', combined)
                    if version_match:
                        return service_name, version_match.group(1)
            
            return product_str.lower(), version_str
        
        def is_version_outdated(service, version):
            """Compare version against latest known version"""
            if service not in latest_versions:
                return False
                
            try:
                current_parts = [int(part) for part in version.split('.')]
                latest_parts = [int(part) for part in latest_versions[service]["version"].split('.')]
                
                while len(current_parts) < len(latest_parts):
                    current_parts.append(0)
                while len(latest_parts) < len(current_parts):
                    latest_parts.append(0)
                
                for i in range(len(current_parts)):
                    if current_parts[i] < latest_parts[i]:
                        return True
                    elif current_parts[i] > latest_parts[i]:
                        return False
                    
                return False
            except ValueError:
                return version < latest_versions[service]["version"]
        
        for tech in self.results["active"].get("web", {}).get("technologies", []):
            service, version = extract_version(tech, "")
            if service in latest_versions and version:
                if is_version_outdated(service, version):
                    service_info = {
                        "host": self.target,
                        "port": "web",
                        "service": tech,
                        "version": version,
                        "recommendation": latest_versions[service]["recommended"]
                    }
                    self.outdated_services["services"].append(service_info)
                    recommendation = f"Update {tech} to {latest_versions[service]['recommended']}"
                    if recommendation not in self.outdated_services["recommendations"]:
                        self.outdated_services["recommendations"].append(recommendation)
        
        for host, host_data in self.results["active"].get("services", {}).items():
            for proto, proto_data in host_data.items():
                for port, port_data in proto_data.items():
                    product = port_data.get("product", "").lower()
                    version = port_data.get("version", "")
                    
                    service, clean_version = extract_version(product, version)
                    
                    if service in latest_versions and clean_version:
                        if is_version_outdated(service, clean_version):
                            display_service = f"{product} {version}".strip()
                            service_info = {
                                "host": host,
                                "port": port,
                                "service": display_service,
                                "version": clean_version,
                                "recommendation": latest_versions[service]["recommended"]
                            }
                            
                            self.outdated_services["services"].append(service_info)
                            
                            recommendation = f"Update {display_service} to {latest_versions[service]['recommended']}"
                            if recommendation not in self.outdated_services["recommendations"]:
                                self.outdated_services["recommendations"].append(recommendation)
        
        self.results["active"]["outdated_services"] = self.outdated_services
        
        print(f"[+] Found {len(self.outdated_services['services'])} outdated services")
        return self.outdated_services
    
    def _parse_nmap_xml(self, xml_data):
        """Parse nmap XML output to extract port and service information"""
        results = {}
        
        host_blocks = re.findall(r"<host.*?</host>", xml_data, re.DOTALL)
        
        for host_block in host_blocks:
            addr_match = re.search(r'<address addr="([^"]*)"', host_block)
            if addr_match:
                host = addr_match.group(1)
                results[host] = {"tcp": {}}
                
                port_blocks = re.findall(r"<port.*?</port>", host_block, re.DOTALL)
                for port_block in port_blocks:
                    port_match = re.search(r'portid="([^"]*)"', port_block)
                    proto_match = re.search(r'protocol="([^"]*)"', port_block)
                    state_match = re.search(r'<state state="([^"]*)"', port_block)
                    service_match = re.search(r'<service name="([^"]*)"', port_block)
                    product_match = re.search(r'product="([^"]*)"', port_block)
                    version_match = re.search(r'version="([^"]*)"', port_block)
                    
                    if port_match and proto_match and state_match:
                        port = port_match.group(1)
                        proto = proto_match.group(1)
                        
                        if proto not in results[host]:
                            results[host][proto] = {}
                            
                        results[host][proto][port] = {
                            "port": port,
                            "state": state_match.group(1),
                            "service": service_match.group(1) if service_match else "unknown"
                        }
                        
                        if product_match:
                            results[host][proto][port]["product"] = product_match.group(1)
                        if version_match:
                            results[host][proto][port]["version"] = version_match.group(1)
                        
                        script_blocks = re.findall(r"<script.*?</script>", port_block, re.DOTALL)
                        if script_blocks:
                            scripts = {}
                            for script_block in script_blocks:
                                script_id_match = re.search(r'id="([^"]*)"', script_block)
                                script_output_match = re.search(r'output="([^"]*)"', script_block)
                                if script_id_match and script_output_match:
                                    scripts[script_id_match.group(1)] = script_output_match.group(1)
                            
                            if scripts:
                                results[host][proto][port]["scripts"] = scripts
        
        return results
    
    def web_scan(self):
        """Perform basic web scanning"""
        print(f"[*] Performing web scan on {self.target}")
        results = {
            "headers": {},
            "technologies": [],
            "interesting_files": []
        }
        
        for port in [80, 443, 8080, 8443]:
            protocol = "https" if port in [443, 8443] else "http"
            url = f"{protocol}://{self.target}:{port}"
            
            try:
                response = requests.get(url, timeout=10, verify=False)
                
                results["headers"][f"{url}"] = dict(response.headers)
                
                if "X-Powered-By" in response.headers:
                    results["technologies"].append(response.headers["X-Powered-By"])
                if "Server" in response.headers:
                    results["technologies"].append(response.headers["Server"])
                
                try:
                    robots = requests.get(f"{url}/robots.txt", timeout=5, verify=False)
                    if robots.status_code == 200:
                        results["interesting_files"].append({
                            "url": f"{url}/robots.txt",
                            "content": robots.text[:500] 
                        })
                except:
                    pass
                
                try:
                    sitemap = requests.get(f"{url}/sitemap.xml", timeout=5, verify=False)
                    if sitemap.status_code == 200:
                        results["interesting_files"].append({
                            "url": f"{url}/sitemap.xml",
                            "content": "Sitemap found"
                        })
                except:
                    pass
                
            except requests.exceptions.RequestException:
                continue
                
        return results
    
    def generate_report(self, format="all"):
        """Generate report in specified format (json, excel, or all)"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"{self.output_dir}/{self.target.replace('.', '_')}_{timestamp}"
        
        if format in ["json", "all"]:
            self._generate_json_report(f"{base_filename}.json")
            
        if format in ["excel", "all"]:
            self._generate_excel_report(f"{base_filename}.xlsx")
            
        if format in ["md", "all"]:
            self._generate_md_report(f"{base_filename}.md")
            
        print(f"[+] Reports generated in {self.output_dir} directory")
    
    def _generate_json_report(self, filename):
        """Generate JSON report"""
        with open(filename, "w") as f:
            json.dump(self.results, f, indent=4)
    
    def _generate_excel_report(self, filename):
        """Generate Excel report"""
        wb = Workbook()
        
        summary = wb.active
        summary.title = "Summary"
        summary['A1'] = "APTES Reconnaissance Report"
        summary['A1'].font = Font(bold=True, size=14)
        summary['A3'] = "Target:"
        summary['B3'] = self.target
        summary['A4'] = "Timestamp:"
        summary['B4'] = self.results["timestamp"]
        
        passive = wb.create_sheet("Passive Recon")
        passive['A1'] = "Passive Reconnaissance Results"
        passive['A1'].font = Font(bold=True, size=14)
        
        row = 3
        passive[f'A{row}'] = "DNS Records"
        passive[f'A{row}'].font = Font(bold=True)
        row += 1
        for record_type, records in self.results["passive"].get("dns", {}).items():
            passive[f'A{row}'] = record_type
            passive[f'B{row}'] = ", ".join(records) if records else "None found"
            row += 1
        
        row += 1
        passive[f'A{row}'] = "Subdomains"
        passive[f'A{row}'].font = Font(bold=True)
        row += 1
        for subdomain in self.results["passive"].get("subdomains", []):
            passive[f'A{row}'] = subdomain
            row += 1
        
        active = wb.create_sheet("Active Recon")
        active['A1'] = "Active Reconnaissance Results"
        active['A1'].font = Font(bold=True, size=14)
        
        active['A3'] = "Open Ports and Services"
        active['A3'].font = Font(bold=True)
        active['A4'] = "IP"
        active['B4'] = "Protocol"
        active['C4'] = "Port"
        active['D4'] = "State"
        active['E4'] = "Service"
        active['F4'] = "Version"
        
        row = 5
        for host, host_data in self.results["active"].get("ports", {}).items():
            for proto, proto_data in host_data.items():
                for port, port_data in proto_data.items():
                    active[f'A{row}'] = host
                    active[f'B{row}'] = proto
                    active[f'C{row}'] = port
                    active[f'D{row}'] = port_data.get("state", "")
                    active[f'E{row}'] = port_data.get("service", "")
                    active[f'F{row}'] = port_data.get("version", "") if port_data.get("product") and port_data.get("version") else ""
                    row += 1
        
        row += 2
        active[f'A{row}'] = "Web Technologies"
        active[f'A{row}'].font = Font(bold=True)
        row += 1
        for tech in self.results["active"].get("web", {}).get("technologies", []):
            active[f'A{row}'] = tech
            row += 1
        
        row += 2
        active[f'A{row}'] = "Outdated Services"
        active[f'A{row}'].font = Font(bold=True)
        row += 1
        active[f'A{row}'] = "Host"
        active[f'B{row}'] = "Port"
        active[f'C{row}'] = "Service"
        active[f'D{row}'] = "Version"
        active[f'E{row}'] = "Recommendation"
        row += 1
        
        for service in self.results["active"].get("outdated_services", {}).get("services", []):
            active[f'A{row}'] = service.get("host", "")
            active[f'B{row}'] = service.get("port", "")
            active[f'C{row}'] = service.get("service", "")
            active[f'D{row}'] = service.get("version", "")
            active[f'E{row}'] = service.get("recommendation", "")
            row += 1
            
        vuln_sheet = wb.create_sheet("Vulnerabilities")
        vuln_sheet['A1'] = "Vulnerability Scan Results"
        vuln_sheet['A1'].font = Font(bold=True, size=14)
        vuln_sheet['A3'] = "Port"
        vuln_sheet['B3'] = "Service"
        vuln_sheet['C3'] = "Vulnerability"
        vuln_sheet['D3'] = "Details"
        
        row = 4
        for vuln in self.results["active"].get("vulnerabilities", {}).get("vulnerabilities", []):
            vuln_sheet[f'A{row}'] = vuln.get("port", "")
            vuln_sheet[f'B{row}'] = vuln.get("service", "")
            vuln_sheet[f'C{row}'] = vuln.get("vulnerability", "")
            vuln_sheet[f'D{row}'] = vuln.get("details", "")[:32767]
            row += 1
        
        wb.save(filename)
    
    def _generate_md_report(self, filename):
        """Generate Markdown report"""
        with open(filename, "w") as f:
            f.write(f"# APTES Reconnaissance Report\n\n")
            f.write(f"## Target: {self.target}\n")
            f.write(f"## Timestamp: {self.results['timestamp']}\n\n")
            
            f.write(f"## Passive Reconnaissance\n\n")
            
            f.write(f"### DNS Records\n\n")
            for record_type, records in self.results["passive"].get("dns", {}).items():
                f.write(f"#### {record_type} Records\n")
                if records:
                    for record in records:
                        f.write(f"- {record}\n")
                else:
                    f.write("- None found\n")
            f.write("\n")
            
            f.write(f"### Subdomains\n\n")
            for subdomain in self.results["passive"].get("subdomains", []):
                f.write(f"- {subdomain}\n")
            f.write("\n")
            
            f.write(f"### SSL Certificate Information\n\n")
            for key, value in self.results["passive"].get("ssl_info", {}).items():
                f.write(f"- **{key}**: {value}\n")
            f.write("\n")
            
            f.write(f"## Active Reconnaissance\n\n")
            
            f.write(f"### Open Ports and Services\n\n")
            f.write(f"| IP | Protocol | Port | State | Service | Version |\n")
            f.write(f"|---|---|---|---|---|---|\n")
            for host, host_data in self.results["active"].get("ports", {}).items():
                for proto, proto_data in host_data.items():
                    for port, port_data in proto_data.items():
                        state = port_data.get("state", "")
                        service = port_data.get("service", "")
                        product = port_data.get("product", "")
                        version = port_data.get("version", "")
                        combined_version = f"{product} {version}".strip()
                        f.write(f"| {host} | {proto} | {port} | {state} | {service} | {combined_version} |\n")
            f.write("\n")
            
            f.write(f"### Web Technologies\n\n")
            for tech in self.results["active"].get("web", {}).get("technologies", []):
                f.write(f"- {tech}\n")
            f.write("\n")
            
            f.write(f"### Web Headers\n\n")
            for url, headers in self.results["active"].get("web", {}).get("headers", {}).items():
                f.write(f"#### {url}\n\n")
                for header, value in headers.items():
                    f.write(f"- **{header}**: {value}\n")
            f.write("\n")
            
            f.write(f"### Interesting Files\n\n")
            for file in self.results["active"].get("web", {}).get("interesting_files", []):
                f.write(f"#### {file.get('url')}\n")
                f.write(f"```\n{file.get('content')}\n```\n\n")
            
            f.write(f"### Outdated Services\n\n")
            f.write(f"| Host | Port | Service | Version | Recommendation |\n")
            f.write(f"|---|---|---|---|---|\n")
            for service in self.results["active"].get("outdated_services", {}).get("services", []):
                host = service.get("host", "")
                port = service.get("port", "")
                service_name = service.get("service", "")
                version = service.get("version", "")
                recommendation = service.get("recommendation", "")
                f.write(f"| {host} | {port} | {service_name} | {version} | {recommendation} |\n")
            f.write("\n")
            
            f.write(f"#### Update Recommendations\n\n")
            for recommendation in self.results["active"].get("outdated_services", {}).get("recommendations", []):
                f.write(f"- {recommendation}\n")
            f.write("\n")
                
            f.write(f"## Vulnerability Scan Results\n\n")
            f.write(f"| Port | Service | Vulnerability | Details |\n")
            f.write(f"|---|---|---|---|\n")
            for vuln in self.results["active"].get("vulnerabilities", {}).get("vulnerabilities", []):
                port = vuln.get("port", "")
                service = vuln.get("service", "")
                vulnerability = vuln.get("vulnerability", "")
                details = vuln.get("details", "").replace("\n", " ").replace("|", "\\|")[:200] + "..."
                f.write(f"| {port} | {service} | {vulnerability} | {details} |\n")
            f.write("\n")
